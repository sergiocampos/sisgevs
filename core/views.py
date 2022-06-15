
#from asyncio.windows_events import NULL
import csv
import json
import os
import string
import urllib
from datetime import datetime
from inspect import Attribute
from io import BytesIO
from random import choice

import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import (authenticate, get_user_model, login, logout,
                                 update_session_auth_hash)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, UserCreationForm
from django.contrib.auth.hashers import check_password, make_password
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models.expressions import OrderBy, Value
from django.http import HttpResponse, HttpResponseRedirect
from django.http.response import JsonResponse
from django.shortcuts import redirect, render
from django.utils.functional import empty
from django.utils.translation import gettext as _
from django.views.decorators.csrf import csrf_protect
from genericpath import exists
from openpyxl import Workbook
from pandas.core.frame import DataFrame

from .models import *

# Create your views here.

def login_page(request):
	storage = messages.get_messages(request)
	storage.used = True
	return render(request, 'login_page.html')


def teste(request):
	return render(request, 'base1.html')
	
@csrf_protect
def login_submit(request):
	if request.POST:
		username = request.POST.get('username')
		password = request.POST.get('password')

		user = authenticate(username = username, password = password)

		
		''' Begin reCAPTCHA validation '''
		recaptcha_response = request.POST.get('g-recaptcha-response')
		url = 'https://www.google.com/recaptcha/api/siteverify'
		values = {
				'secret': settings.GOOGLE_RECAPTCHA_SECRET_KEY,
				'response': recaptcha_response
			}
		data = urllib.parse.urlencode(values).encode()
		req =  urllib.request.Request(url, data=data)
		response = urllib.request.urlopen(req)
		result = json.loads(response.read().decode())
		''' End reCAPTCHA validation '''
		if result['success']:
			if user is not None:
				login(request, user)
				#return redirect('/all_forms/')
				return redirect('principal')
			else:
				messages.error(request, 'Senha ou usuário inválidos.')
				return render(request, 'login_page.html')

			#form.save()
			messages.success(request, 'New comment added with success!')

		else:
			messages.error(request, 'CAPTCHA inválido. Tente novamente.')
		
			#messages.error(request, 'Usuário e/ou senha inválido!')
	return redirect('/login/')


@login_required(login_url='/login/')
def logout_user(request):
    logout(request)
    return redirect('/')


@login_required
def change_password(request):
	if request.method == 'POST':
		form = PasswordChangeForm(request.user, request.POST)
		if form.is_valid():
			user = form.save()
			update_session_auth_hash(request, user)
			messages.success(request, _('Your password was successfully updated!'))
			return redirect('/all_forms/')
		else:
			messages.error(request, _('Please correct the error below.'))
	else:
		form = PasswordChangeForm(request.user)
	return render(request, 'change_password.html', {'form': form})


@login_required(login_url='/login/')
def set_login_page(request):
	return redirect('index')



@login_required(login_url='/login/')
def index(request):
	return render(request, 'index.html')



@login_required(login_url='/login/')
def dados_user(request):
	return render(request, 'dados_user.html')


@login_required(login_url='/login/')
def main(request):
	return render(request, 'main.html')



@login_required(login_url='/login/')
def all_forms(request):
	# notificadores = ['autocadastro', 'admin', 'municipal', 'gerencia_executiva', 'gerencia_operacional', 'chefia_nucleo', 'area_tecnica']
	if request.user.funcao != 'gerencia_regional':
		return render(request, 'all_forms.html')
	else:
		return redirect('/dados_user/')

########################  View de localização de uma notificação ##################################
@login_required(login_url='/login/')
def localizar_paciente_nome(request):
	return render(request, 'localizar_paciente_nome.html')


@login_required(login_url='/login/')
def set_localizar_paciente_nome(request):
	nome = request.POST.get('nome')
	request.session['nome'] = nome
	return redirect('search_paciente_nome')

@login_required(login_url='/login/')
def search_paciente_nome(request):
	nome = request.session['nome']
	caso_all_result = CasoEsporotricose.objects.filter(nome_paciente__icontains=nome)

	paginator = Paginator(caso_all_result, 6)
	page = request.GET.get('page')
	regs = paginator.get_page(page)

	return render(request, 'resultado_search_caso_nome.html', {'regs':regs})


# Localizar Paciente por Data da Coleta
@login_required(login_url='/login/')
def localizar_paciente_data_coleta(request):
	return render(request, 'localizar_paciente_data_coleta.html')


@login_required(login_url='/login/')
def set_localizar_paciente_data_coleta(request):
	data = request.POST.get('data')
	caso_all_result = CasoEsporotricose.objects.filter(data_coleta1__range=[data, data])
		
	return render(request, 'resultado_search_caso_data_coleta.html', {'caso_all_result':caso_all_result})

@login_required(login_url='/login/')
def csv_localizar_paciente_data_coleta(request):
	date = request.POST.get('data_coleta')
	caso_all_result = CasoEsporotricose.objects.filter(data_coleta1__range=[date, date])
	
	data = []
	for item in caso_all_result:
		data.append([item.tipo_notificacao, item.agravo_doenca, item.codigo_cib10, item.data_notificacao, item.estado, item.municipio, item.codigo_ibge, item.data_primeiros_sintomas, item.unidade_saude, item.nome_paciente, item.data_nascimento_paciente, item.idade_paciente, item.sexo_paciente, item.paciente_gestante, item.raca_paciente, item.escolaridade_paciente, item.cartao_sus_paciente, item.nome_mae_paciente, item.cep_residencia, item.uf_residencia, item.municipio_residencia, item.bairro_residencia, item.codigo_ibge_residencia, item.rua_residencia, item.numero_residencia, item.complemento_residencia, item.distrito_residencia, item.ponto_referencia_residencia, item.telefone_residencia, item.zona_residencia, item.pais_residencia, item.data_investigacao, item.ocupacao, item.ambientes_frequentados, item.animais_que_teve_contato, item.natureza_contato_animais, item.relacao_animal_doente, item.exerce_atividade_contato_plantas, item.historico_contato_material, item.presenca_lesao_pele, item.natureza_lesao, item.local_lesao, item.diagnostico_forma_extrac_doenca, item.localizacao_forma_extrac_doenca, item.houve_coleta_material, item.data_coleta1, item.numero_gal1, item.data_coleta2, item.numero_gal2, item.data_coleta3, item.numero_gal3, item.resultado_isolamento, item.agente, item.histopatologia, item.data_resultado_exame1, item.descricao_exame_1, item.resultado_exame1, item.data_resultado_exame2, item.descricao_exame_2, item.resultado_exame2, item.data_resultado_exame3, item.descricao_exame_3, item.resultado_exame3, item.data_inicio_tratamento1, item.droga_administrada1, item.esquema_terapeutico1, item.data_inicio_tratamento2, item.droga_administrada2, item.esquema_terapeutico2, item.data_inicio_tratamento3, item.droga_administrada3, item.esquema_terapeutico3, item.hospitalizacao, item.data_internacao, item.data_da_alta, item.uf_hospitalizacao, item.municipio_hospitalizacao, item.codigo_ibge_hospitalizacao, item.nome_hospital_hospitalizacao, item.classificacao_final, item.criterio_confirmacao, item.caso_autoctone_municipio_residencia, item.uf_caso_autoctone, item.pais_caso_autoctone, item.municipio_caso_autoctone, item.codigo_ibge_caso_autoctone, item.distrito_caso_autoctone, item.bairro_caso_autoctone, item.area_provavel_infeccao_caso_autoctone, item.ambiente_infeccao_caso_autoctone, item.doenca_rel_trabalho_caso_autoctone, item.evolucao_caso, item.data_obito, item.data_encerramento, item.observacao, item.nome_investigador, item.funcao_investigador, item.email_investigador, item.telefone_investigador, item.conselho_classe_investigador, item.responsavel_pelas_informacoes_id, item.unidade_saude_outro, item.gerencia_id])
	data = pd.DataFrame(data, columns=['Tipo de Notificação', 'Agravo/Doença', 'Codigo CID-10', 'Data de Notificação', 'Estado', 'Municipio', 'Codigo IBGE', 'Data dos Primeiros Sintomas', 'Unidade de Saúde', 'Nome do Paciente', 'Data de Nascimento do Paciente', 'Idade do Paciente', 'Sexo do Paciente', 'Paciente Gestante', 'Raça do Paciente', 'Escolaridade do Paciente', 'Cartão SUS do Paciente', 'Nome da Mãe do Paciente', 'CEP da Residência', 'UF da Residência', 'Municipio da Residência', 'Bairro da Residência', 'Codigo IBGE da Residência', 'Rua da Residência', 'Numero_da Residência', 'Complemento da Residência', 'Distrito da Residência', 'Ponto de Referência da Residência', 'Telefone da Residência', 'Zona da Residência', 'País da Residência', 'Data de Investigação', 'Ocupação', 'Ambientes Frequentados', 'Animais que Teve Contato', 'Natureza do Contato com os Animais', 'Relação com o Animal Doente', 'Exerce Atividade Contato Plantas', 'Histórico Contato Material', 'Presença de Lesão na Pele', 'Natureza da Lesão', 'Local da Lesão', 'Diagnóstico Forma Extrac Doença', 'Localização Forma Extrac Doença', 'Houve Coleta de Material', 'Data da Coleta 1', 'Numero Gal 1', 'Data da Coleta 2', 'Numero Gal 2', 'Data da Coleta 3', 'Numero Gal 3', 'Resultado de Isolamento', 'Agente', 'Histopatologia', 'Data do Restultado EXAME 1', 'Descrição EXAME 1', 'Resultado EXAME 1', 'Data do Resultado EXAME 2', 'Descrição EXAME 2', 'Resultado EXAME 2', 'Data do Resultado EXAME 3', 'Descrição EXAME 3', 'Resultado EXAME 3', 'Data de Início do Tratamento 1', 'Droga Administrada 1', 'Esquema Terapeutico 1', 'Data de Início do Tratamento 2', 'Droga Administrada 2', 'Esquema Terapeutico 2', 'Data de Início do Tratamento 3', 'Droga Administrada 3', 'Esquema Terapeutico 3', 'Hospitalizacão', 'Data de Internação', 'Data da Alta', 'UF da Hospitalização', 'Município da Hospitalização', 'Codigo IBGE da Hospitalização', 'Nome do Hospital da Hospitalização', 'Classificação Final', 'Critério de Confirmação', 'Caso Autoctone Municipio Residência', 'UF Caso Autoctone', 'País Caso Autoctone', 'Município Caso Autoctone', 'Codigo IBGE Caso Autoctone', 'Distrito Caso Autoctone', 'Bairro Caso Autoctone', 'Área Provável de Infecção Caso Autoctone', 'Ambiente Infecção Caso Autoctone', 'Doença Rel Trabalho Caso Autoctone', 'Evolução do Caso', 'Data do Óbito', 'Data de Encerramento', 'Observação', 'Nome do Investigador', 'Função do Investigador', 'Email do Investigador', 'Telefone do Investigador', 'Conselho Classe Investigador', 'Responsável Pelas Informações ID', 'Unidade de Saude Outro', 'Gerencia ID'])
	
	response = HttpResponse(content_type = "text/csv")
	response['Content-Disposition'] = 'attachment; filename=casos_esporotricose.csv'
	data.to_csv(response, index=False)
		
	return response
# Localizar Paciente por Data da Coleta


###########View Renderiza a ficha para impressão######################################################
@login_required(login_url='/login/')
def download_ficha(request):
	file_path = os.path.join(settings.MEDIA_ROOT, 'ficha_esporotricose_humana.pdf')
	if os.path.exists(file_path):
		with open(file_path, 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/pdf")
			response['Content-Disposition'] = 'inline; filename' + os.path.basename(file_path)
			return response



###########View Renderiza o dicionario de dados#######################################################
@login_required(login_url='/login/')
def download_dicionario_dados(request):
	file_path = os.path.join(settings.MEDIA_ROOT, 'dicionario_de_dados_Esporotricose_Humana.pdf')
	if os.path.exists(file_path):
		with open(file_path, 'rb') as fh:
			response = HttpResponse(fh.read(), content_type="application/pdf")
			response['Content-Disposition'] = 'inline; filename' + os.path.basename(file_path)
			return response



@login_required(login_url='/login/')
def caso_view(request, id):
	registro = CasoEsporotricose.objects.get(id=id)
	municipio_id = registro.municipio
	if municipio_id == None or municipio_id == '':
		return render(request, 'caso_view.html', {'registro':registro})
	else:
		municipio = Municipio.objects.get(id=municipio_id)
		ibge = municipio.ibge
		return render(request, 'caso_view.html', {'registro':registro, 'municipio':municipio, 'ibge':ibge})
		


@login_required(login_url='/login/')
def caso_view_detail(request, id):
	registro = CasoEsporotricose.objects.get(id=id)
	uf_residencia = registro.uf_residencia
	registro_cidade_residencia_str = registro.municipio_residencia
	try:
		uf = Estado.objects.get(id=uf_residencia)
	except:
		uf = None
	
	if registro.municipio != None:
		municipio_id = registro.municipio
		municipio = Municipio.objects.get(id=municipio_id)
	else:
		municipio = None
	
	

	return render(request, 'caso_view_detail.html', {'registro':registro, 'municipio':municipio, 'uf_residencia':uf})


@login_required(login_url='/login/')
def caso_esporotricose_create(request):
	estados = Estado.objects.all().order_by('nome')
	municipios = Municipio.objects.all().order_by('nome')
	unidades_saude_municipio = UnidadeSaude.objects.filter(municipio=request.user.municipio)
	unidades_saude = []
	codigos_ibge = []
	return render(request, 'caso_esporotricose_create.html', {'municipios':municipios, 'unidades_saude':unidades_saude, 
		'codigos_ibge':codigos_ibge, 'estados':estados, 'unidades_saude_municipio':unidades_saude_municipio, 'estados':estados})

#############################views ajax dados gerais#####################################

@login_required(login_url='/login/')
def ajax_load_unidadesaude(request):
	municipio_id = request.GET.get('municipio_id')
	#cod_ibge = JoinMunicipioIbgeUnidadeSaude.objects.filter(municipio=municipio).all()
	cod_ibge = UnidadeSaude.objects.filter(municipio_id=municipio_id).all().order_by('nome')
	
	return render(request, 'unidades_saude_ajax.html', {'cod_ibge':cod_ibge})


@login_required(login_url='/login/')
def ajax_load_ibge(request):
	municipio_id = request.GET.get('municipio_id')
	codigo = CodigoIbge.objects.filter(municipio_id=municipio_id).values()

	return render(request, 'ibge_ajax.html', {'codigo':codigo})


#############################view do ajax para hospitalização#################################
@login_required(login_url='/login/')
def ajax_hospitalizacao(request):
	municipio_id = request.GET.get('municipio_id')
	cod_ibge = UnidadeSaude.objects.filter(municipio_id=municipio_id).all().exclude(nome__contains='USF').order_by('nome')
	
	return render(request, 'hospitalizacao_ajax.html', {'cod_ibge':cod_ibge})


@login_required(login_url='/login/')
def ajax_hospitalizacao_ibge(request):
	municipio_id = request.GET.get('municipio_id')
	codigo = CodigoIbge.objects.filter(municipio_id=municipio_id).values()

	return render(request, 'hospitalizacaoibge_ajax.html', {'codigo':codigo})


########################## view para caso autoctone ##########################################
@login_required(login_url='/login/')
def ajax_autoctone_uf(request):
	uf_id = request.GET.get('municipio_id')
	#cod_ibge = JoinMunicipioIbgeUnidadeSaude.objects.filter(municipio=municipio).all()
	municipio = Municipios.objects.filter(uf_id=uf_id).all()
	
	return render(request, 'municipios_estado_ajax.html', {'municipio':municipio})


@login_required(login_url='/login/')
def ajax_autoctone_municipio(request):
	id = request.GET.get('municipio_id')
	#cod_ibge = JoinMunicipioIbgeUnidadeSaude.objects.filter(municipio=municipio).all()
	municipio = Municipios.objects.get(id=id)
	ibge = municipio.ibge
	return render(request, 'municipios_ibge_ajax.html', {'ibge':ibge})


@login_required(login_url='/login/')
def ajax_autoctone_distrito(request):
	municipio_id = request.GET.get('municipio_id')
	#cod_ibge = JoinMunicipioIbgeUnidadeSaude.objects.filter(municipio=municipio).all()
	distrito = Distrito.objects.filter(municipio_id=municipio_id)
	return render(request, 'municipios_distrito_ajax.html', {'distrito':distrito})

############################## View para dados residencia#########################################

@login_required(login_url='/login/')
def ajax_dados_residencia(request):
	estado_id = request.GET.get('uf_dados_residencia_id')
	if estado_id == '12':
		municipios = Municipio.objects.all()
	else:
		municipios = Municipios.objects.filter(uf_id=estado_id)

	return render(request, 'estado_municipios_ajax.html', {'municipios':municipios})



@login_required(login_url='/login/')
def ajax_ibge_municipio_residencia(request):
	municipio_estado_id = request.GET.get('municipios_estado_id')
	uf_estado_id = request.GET.get('uf_estado_id')
	
	if uf_estado_id == '12':
		municipio = Municipio.objects.get(id=municipio_estado_id)
	else:
		municipio = Municipios.objects.get(id=municipio_estado_id)
	ibge = municipio.ibge
	
	
	return render(request, 'municipio_ibge_ajax.html', {'ibge':ibge})


#################################### views ajax para edição de endereço################################
@login_required(login_url='/login/')
def ajax_edicao_uf_cidades(request):
	estado_id = request.GET.get('uf_edit_endereco_id')
	if estado_id == '12':
		municipios = Municipio.objects.all()
	else:
		municipios = Municipios.objects.filter(uf_id=estado_id)

	return render(request, 'edit_estado_municipios_ajax.html', {'municipios':municipios})
#######################################################################################################

@login_required(login_url='/login/')
def casos_cancelados(request):
	municipios = Municipio.objects.all()
	if request.user.funcao == 'admin':
		registros = CasoEsporotricose.objects.filter(status_caso='Cancelado').order_by('-data_notificacao')
		return render(request, 'casos_cancelados.html', {'regs':registros, 'municipios':municipios})
	else:
		return redirect('all_forms')

@login_required(login_url='/login/')
def my_datas(request):
	
	municipios = Municipio.objects.all()
	#municipio_nome = municipio_user.nome
	if request.user.funcao == 'admin':
		registros = CasoEsporotricose.objects.all().order_by('-data_notificacao').exclude(status_caso='Cancelado')
		
		paginator = Paginator(registros, 6)
		page = request.GET.get('page')
		regs = paginator.get_page(page)

		return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})

	elif request.user.funcao == 'gerencia_executiva':
		registros = CasoEsporotricose.objects.all().order_by('-data_notificacao').exclude(status_caso='Cancelado')
		
		paginator = Paginator(registros, 6)
		page = request.GET.get('page')
		regs = paginator.get_page(page)
		
		return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})

	elif request.user.funcao == 'gerencia_operacional':
		if request.user.has_perm('core.acessa_esporotricose'):
			user_gerencia_operacional = request.user.gerencia_operacional
			registros = CasoEsporotricose.objects.all().order_by('-data_notificacao').exclude(status_caso='Cancelado')
			
			paginator = Paginator(registros, 6)
			page = request.GET.get('page')
			regs = paginator.get_page(page)
		
			return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})
		

	elif request.user.funcao == 'chefia_nucleo':
		if request.user.has_perm('core.acessa_esporotricose'):
			user_gerencia_operacional = request.user.gerencia_operacional
			user_nucleo = request.user.nucleo
			registros = CasoEsporotricose.objects.all().order_by('-data_notificacao').exclude(status_caso='Cancelado')
			
			paginator = Paginator(registros, 6)
			page = request.GET.get('page')
			regs = paginator.get_page(page)
			
			return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})	

	elif request.user.funcao == 'area_tecnica':
		if request.user.has_perm('core.acessa_esporotricose'):
			user_gerencia_operacional = request.user.gerencia_operacional
			user_nucleo = request.user.nucleo
			user_area_tecnica = request.user.area_tecnica
			registros = CasoEsporotricose.objects.all().order_by('-data_notificacao').exclude(status_caso='Cancelado')
			
			paginator = Paginator(registros, 6)
			page = request.GET.get('page')
			regs = paginator.get_page(page)
			
			return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})	

	elif request.user.funcao == 'gerencia_regional':
		
		user_gerencia_regional = Municipio.objects.get(id=request.user.municipio_id).gerencia_id
		registros = CasoEsporotricose.objects.filter(gerencia_id=user_gerencia_regional).order_by('-data_notificacao').exclude(status_caso='Cancelado')

		paginator = Paginator(registros, 6)
		page = request.GET.get('page')
		regs = paginator.get_page(page)
		
		return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})

	elif request.user.funcao == 'municipal':
		
		user_id = request.user.id
		user_municipio_id = request.user.municipio_id
		municipio_user = request.user.municipio
		user_municipio_nome = str(Municipio.objects.filter(id=user_municipio_id)[0])
		user_municipio_nome_upper = str(Municipio.objects.filter(id=user_municipio_id)[0]).upper()
		
		registros = CasoEsporotricose.objects.filter(
			Q(municipio_residencia=user_municipio_id) | 
			Q(municipio_residencia=user_municipio_nome) | 
			Q(municipio_residencia=user_municipio_nome_upper) | 
			Q(responsavel_pelas_informacoes_id=user_id)
			).order_by('-data_notificacao').exclude(status_caso='Cancelado')
		
		paginator = Paginator(registros, 6)
		page = request.GET.get('page')
		regs = paginator.get_page(page)

		return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})

	elif request.user.funcao == 'autocadastro':
		autocadastro_id = request.user.id
		municipio_user = request.user.municipio

		registros = CasoEsporotricose.objects.filter(responsavel_pelas_informacoes_id=autocadastro_id).order_by('-data_notificacao').exclude(status_caso='Cancelado')
		#registros = CasoEsporotricose.objects.filter(municipio_residencia=municipio_user).order_by('-data_notificacao')

		
		paginator = Paginator(registros, 6)
		page = request.GET.get('page')
		regs = paginator.get_page(page)
		
		return render(request, 'my_datas.html', {'regs':registros, 'municipios':municipios})

	return redirect('all_forms')

@login_required(login_url='/login/')
def cancelar_caso_esporotricose(request):
	caso_id = request.GET.get('id')
	redirect_url = request.GET.get('url_redirect')
	registro = CasoEsporotricose.objects.filter(id=caso_id).values()
	if registro[0]['status_caso'] == 'Cancelado':
		CasoEsporotricose.objects.filter(id=caso_id).update(status_caso=None)
	else:		
		CasoEsporotricose.objects.filter(id=caso_id).update(status_caso='Cancelado')	
	return redirect(redirect_url)

@login_required(login_url='/login/')
def criar_perfil_municipal(request):
	if request.method == 'GET':
		data = request.GET.get('nome')
		if data == None:
			municipios = Municipio.objects.all().order_by('nome')
			municipios = {'municipios':municipios}
			return render(request, 'criar_perfil_municipal.html', municipios)
		
		# Recebendo dados de cadastro.
		else:
			nome = data
			email = request.GET.get('email')
			cpf = request.GET.get('cpf')
			login = request.GET.get('login')
			funcao = request.GET.get('perfil')
			municipio = int(request.GET.get('municipio'))
			gerencia_regional = request.GET.get('gerencia_regional')
			gerencia_operacional = request.GET.get('gerencia_operacional')
			nucleo = request.GET.get('nucleo')
			area_tecnica = request.GET.get('area_tecnica')
			telefone = request.GET.get('telefone')
			
			cpf = str(cpf).replace('.','').replace('-', '')			
			telefone = telefone.replace('(','').replace(')','').replace('.','').replace('-','')			
			nome = str(nome).upper()			

			# Criando senha.
			tamanho = 8
			valores = string.ascii_letters + string.digits
			senha = ''
			for i in range(tamanho):
				senha += choice(valores)
			password = make_password(password=senha, salt=None, hasher='pbkdf2_sha256')
				 		
			User = get_user_model()
			User.objects.create(
				login=login,
				funcao=funcao,
				username=nome,
				cpf=cpf,
				email=email,
				telefone=telefone,
				first_name=nome,
				gerencia_operacional=gerencia_operacional,
				nucleo=nucleo,
				area_tecnica=area_tecnica,
				gerencia_regional=gerencia_regional,
				municipio = Municipio.objects.get(id=municipio),
				password=password
			)
			
			data = {'login':login, 'senha':senha}
			return JsonResponse(data)


@login_required(login_url='/login/')
def checar_login_ajax(request):
	User = get_user_model()
	
	login = request.GET.get('login')
	usuarios = User.objects.filter(login=login)
	data = {'data':len(usuarios)}
	
	return JsonResponse(data)


@login_required(login_url='/login/')
def ficha_caso_esporotricose_preencher(request):
	return render(request, 'ficha_caso_esporotricose_preencher.html')

@login_required(login_url='/login/')
def ficha_caso_esporotricose_preenchido(request):
	return render(request, 'ficha_caso_esporotricose_preenchido.html')




# AJAX GAL
@login_required(login_url='/login/')
def ajax_gal(request):
	n_gal = request.GET.get('n_gal')
	#conn = psycopg2.connect(dbname='yourdb', user='dbuser', password='abcd1234', Host='server', port='5432', sslmode='require')

	# DATA 
	

	# DATA 




	return JsonResponse()
# AJAX GAL





def remove_caso_esporotricose(request, id):
	caso_esporotricose = CasoEsporotricose.objects.get(id=id)
	caso_esporotricose.delete()
	return redirect('my_datas')

@login_required(login_url='/login/')
def set_caso_esporotricose_create(request):
	# Dados do responsável pela criação do caso.
	responsavel_pelas_informacoes = request.user
	responsavel_gerencia_operacional = request.user.gerencia_operacional
	responsavel_nucleo = request.user.nucleo
	responsavel_area_tecnica = request.user.area_tecnica
	responsavel_gerencia_regional = request.user.gerencia_regional
	responsavel_municipio = request.user.municipio

	#Dados Gerais
	tipo_notificacao = request.POST.get('tipo_notificacao')
	agravo_doenca = request.POST.get('agravo_doenca')
	codigo_cib10 = request.POST.get('codigo_cid')
	
	data_notificacao_cap = request.POST.get('data_notificacao')
	
	if data_notificacao_cap == '' or data_notificacao_cap == None:
		data_notificacao = None
	else:
		data_notificacao = datetime.strptime(data_notificacao_cap, '%Y-%m-%d').date()

	
	estado = request.POST.get('uf_notificacao')
	
	municipio_cap = request.POST.get('municipio_notificacao')
	if municipio_cap == '' or municipio_cap == None:
		municipio = None
	else:
		municipio = int(municipio_cap)

	municipio_id = municipio
	municipio_ = Municipio.objects.get(id=municipio_id)


	gerencia = municipio_.gerencia

	codigo_ibge = request.POST.get('codigo_ibge_dados_gerais')
	
	data_primeiros_sintomas_cap = request.POST.get('data_primeiros_sintomas')
	if data_primeiros_sintomas_cap == '' or data_primeiros_sintomas_cap == None:
		data_primeiros_sintomas = None
	else:
		data_primeiros_sintomas = datetime.strptime(data_primeiros_sintomas_cap, '%Y-%m-%d').date()

	
	unidade_saude = request.POST.get('unidade_saude')
	unidade_saude_outro = request.POST.get('unidade_saude_outro')
	
	#notificação individual
	nome_paciente = request.POST.get('nome_paciente').upper()
	
	data_nascimento_paciente_cap = request.POST.get('data_nasc')
	if data_nascimento_paciente_cap == '' or data_nascimento_paciente_cap == None:
		data_nascimento_paciente = None
	else:
		data_nascimento_paciente = datetime.strptime(data_nascimento_paciente_cap, '%Y-%m-%d').date()

	sexo_paciente = request.POST.get('sexo')
	
	idade_paciente_cap = request.POST.get('result')
	if idade_paciente_cap == '' or idade_paciente_cap == None:
		idade_paciente = None
	else:
		idade_paciente = int(idade_paciente_cap)

	paciente_gestante = request.POST.get('gestacao')
	raca_paciente = request.POST.get('raca')
	escolaridade_paciente = request.POST.get('escolaridade')
	cartao_sus_paciente = request.POST.get('cartao_sus')
	nome_mae_paciente = request.POST.get('nome_mae').upper()
	
	#Dados Residencia
	cep_residencia = request.POST.get('cep_residencia')
	uf_residencia = request.POST.get('uf_residencia')
	municipio_residencia = int(request.POST.get('cidade_residencia'))

	if uf_residencia == '12': # Se o uf for PB pega dados no modelo Municipio
		municipio_residencia = Municipio.objects.get(id=municipio_residencia)
	else: # Caso contrário pega os dados no modelo Municipios
		municipio_residencia = Municipios.objects.get(id=municipio_residencia)
	municipio_residencia = municipio_residencia.nome

	bairro_residencia = request.POST.get('bairro_residencia')
	codigo_ibge_residencia = request.POST.get('codigo_ibge_residencia')
	rua_residencia = request.POST.get('rua_residencia')
	numero_residencia = request.POST.get('numero_residencia')
	complemento_residencia = request.POST.get('complemento_residencia')
	distrito_residencia = request.POST.get('distrito_residencia')
	ponto_referencia_residencia = request.POST.get('ponto_referencia_residencia')
	telefone_residencia = request.POST.get('telefone_residencia')
	zona_residencia = request.POST.get('zona_residencia')
	pais_residencia = request.POST.get('pais_residente_fora_pais')
	
	#Antecedentes Epidemiologicos
	data_investigacao_cap = request.POST.get('data_investigacao')
	if data_investigacao_cap == '' or data_investigacao_cap == None:
		data_investigacao = None
	else:
		data_investigacao = datetime.strptime(data_investigacao_cap, '%Y-%m-%d').date()


	ocupacao = request.POST.get('ocupacao')
	ambientes_frequentados = request.POST.getlist('ambientes_frequentados')	
	ambientes_frequentados_outros = request.POST.get('ambientes_frequentados_outros')	
	animais_que_teve_contato = request.POST.getlist('animais_que_teve_contato')	
	animais_que_teve_contato_outros = request.POST.get('animais_que_teve_contato_outros')	
	natureza_contato_animais = request.POST.getlist('natureza_contato_animais')	
	natureza_contato_animais_outros = request.POST.get('natureza_contato_animais_outros')	
	relacao_animal_doente = request.POST.getlist('relacao_animal_doente')	
	relacao_animal_doente_outros = request.POST.get('relacao_animal_doente_outros')	
	exerce_atividade_contato_plantas = request.POST.get('exerc_ativ_contato_plantas')
	historico_contato_material = request.POST.get('lesao_manuseio')
	
	#dados clínicos
	presenca_lesao_pele = request.POST.get('presenca_lesao_pele')
	natureza_lesao = request.POST.getlist('natureza_lesao')
	natureza_lesao_outro = request.POST.get('natureza_lesao_outros')
	local_lesao = request.POST.getlist('local_lesao')
	local_lesao_outro = request.POST.get('local_lesao_outros')
	diagnostico_forma_extrac_doenca = request.POST.get('diagnostico_extracutaneo')
	localizacao_forma_extrac_doenca = request.POST.get('local_extra_cutanea')
	
	#Dados Laboratoriais
	houve_coleta_material = request.POST.get('coleta_exame_labo_sim')
	
	data_coleta1 = request.POST.get('data_coleta_mat_exame1')
	if data_coleta1 == '' or data_coleta1 == None:
		data_coleta1 = None
		numero_gal1 = None
	else:
		data_coleta1 = datetime.strptime(data_coleta1, '%Y-%m-%d').date()
		numero_gal1 = request.POST.get('numero_gal1')

	data_coleta2 = request.POST.get('data_coleta_mat_exame2')
	if data_coleta2 == '' or data_coleta2 == None:
		data_coleta2 = None
		numero_gal2 = None
	else:
		data_coleta2 = datetime.strptime(data_coleta2, '%Y-%m-%d').date()
		numero_gal2 = request.POST.get('numero_gal2')

	data_coleta3 = request.POST.get('data_coleta_mat_exame3')
	if data_coleta3 == '' or data_coleta3 == None:
		data_coleta3 = None
		numero_gal3 = None
	else:
		data_coleta3 = datetime.strptime(data_coleta3, '%Y-%m-%d').date()
		numero_gal3 = request.POST.get('numero_gal3')
	
	resultado_isolamento = request.POST.get('resultado_isolamento')
	agente = request.POST.get('agente_isolamento')
	histopatologia = request.POST.get('histopatologia')
	
	data_exame1_cap = request.POST.get('data_outros_exames_1')
	if data_exame1_cap == '' or data_exame1_cap == None:
		data_resultado_exame1 = None
		descricao_exame_1 = None
		resultado_exame1 = None
	else:
		data_resultado_exame1 = datetime.strptime(data_exame1_cap, '%Y-%m-%d').date()
		descricao_exame_1 = request.POST.get('descricao_outros_exames_1')
		resultado_exame1 = request.POST.get('resultado_outros_exames_1')
	

	
	data_exame2_cap = request.POST.get('data_outros_exames_2')
	if data_exame2_cap == '' or data_exame2_cap == None:
		data_resultado_exame2 = None
		descricao_exame_2 = None
		resultado_exame2 = None
	else:
		data_resultado_exame2 = datetime.strptime(data_exame2_cap, '%Y-%m-%d').date()
		descricao_exame_2 = request.POST.get('descricao_outros_exames_2')
		resultado_exame2 = request.POST.get('resultado_outros_exames_2')

	
	
	data_exame3_cap = request.POST.get('data_outros_exames_3')
	if data_exame3_cap == '' or data_exame3_cap == None:
		data_resultado_exame3 = None
		descricao_exame_3 = None
		resultado_exame3 = None
	else:
		data_resultado_exame3 = datetime.strptime(data_exame3_cap, '%Y-%m-%d').date()
		descricao_exame_3 = request.POST.get('descricao_outros_exames_3')
		resultado_exame3 = request.POST.get('resultado_outros_exames_3')
		
	
	
	#tratamento
	data_inicio_tratamento1_cap = request.POST.get('data_inicio_tratamento')
	if data_inicio_tratamento1_cap == '' or data_inicio_tratamento1_cap == None:
		data_inicio_tratamento1 = None		
	else:
		data_inicio_tratamento1 = datetime.strptime(data_inicio_tratamento1_cap, '%Y-%m-%d').date()
	

	droga_administrada1 = request.POST.get('droga_admin_tratamento')
	esquema_terapeutico1 = request.POST.get('esquema_terapeutico')
	
	data_inicio_tratamento2_cap = request.POST.get('data_inicio_tratamento2')
	if data_inicio_tratamento2_cap == '' or data_inicio_tratamento2_cap == None:
		data_inicio_tratamento2 = None
	else:
		data_inicio_tratamento2 = datetime.strptime(data_inicio_tratamento2_cap, '%Y-%m-%d').date()
	
	
	droga_administrada2 = request.POST.get('droga_admin_tratamento2')
	esquema_terapeutico2 = request.POST.get('esquema_terapeutico2')
	
	data_inicio_tratamento3_cap = request.POST.get('data_inicio_tratamento3')
	if data_inicio_tratamento3_cap == '' or data_inicio_tratamento3_cap == None:
		data_inicio_tratamento3 = None
	else:
		data_inicio_tratamento3 = datetime.strptime(data_inicio_tratamento3_cap, '%Y-%m-%d').date()
	

	droga_administrada3 = request.POST.get('droga_admin_tratamento3')
	esquema_terapeutico3 = request.POST.get('esquema_terapeutico3')
	hospitalizacao = request.POST.get('hospitalizado')
	
	data_internacao_cap = request.POST.get('data_internacao')
	if data_internacao_cap == '' or data_internacao_cap == None:
		data_internacao = None
	else:
		data_internacao = datetime.strptime(data_internacao_cap, '%Y-%m-%d').date()
	

	data_da_alta_cap = request.POST.get('data_alta')
	if data_da_alta_cap == '' or data_da_alta_cap == None:
		data_da_alta = None
	else:
		data_da_alta = datetime.strptime(data_da_alta_cap, '%Y-%m-%d').date()
	

	uf_hospitalizacao = request.POST.get('uf_internacao')
	municipio_hospitalizacao = request.POST.get('municipio_hospital')
	codigo_ibge_hospitalizacao = request.POST.get('codigo_ibge_municipio_hospitalizacao')
	nome_hospital_hospitalizacao = request.POST.get('nome_hospital_hospitalizacao')
	
	#conclusao
	classificacao_final = request.POST.get('classificacao')
	criterio_confirmacao = request.POST.get('criterio')
	caso_autoctone_municipio_residencia = request.POST.get('autoctone')
	uf_caso_autoctone = request.POST.get('uf_caso_autoctone_municipio')
	pais_caso_autoctone = request.POST.get('pais_autoctone_municipio')
	municipio_caso_autoctone = request.POST.get('municipio_autoctone')
	codigo_ibge_caso_autoctone = request.POST.get('codigo_ibge_municipio_autoctone')
	distrito_caso_autoctone = request.POST.get('distrito_autoctone')
	bairro_caso_autoctone = request.POST.get('bairro_autoctone')
	area_provavel_infeccao_caso_autoctone = request.POST.get('area_provavel_infeccao')
	ambiente_infeccao_caso_autoctone = request.POST.get('ambiente_infeccao_autoctone')
	doenca_rel_trabalho_caso_autoctone = request.POST.get('doenca_rel_trabalho_autoctone')
	evolucao_caso = request.POST.get('evolucao_caso_autoctone')
	status_caso = request.POST.get('status_caso')

	data_obito_cap = request.POST.get('data_obito')
	if data_obito_cap == '' or data_obito_cap == None:
		data_obito = None
	else:
		data_obito = datetime.strptime(data_obito_cap, '%Y-%m-%d').date()
	

	data_encerramento_cap = request.POST.get('data_encerramento')
	if data_encerramento_cap == '' or data_encerramento_cap == None:
		data_encerramento = None		
	else:
		data_encerramento = datetime.strptime(data_encerramento_cap, '%Y-%m-%d').date()
	
	#observação
	observacao = request.POST.get('observacao')
	
	#investigador
	nome_investigador = request.POST.get('nome_investigador')
	funcao_investigador = request.POST.get('funcao_investigador')
	email_investigador = request.POST.get('email_investigador')
	telefone_investigador = request.POST.get('telefone_investigador')
	conselho_classe_investigador = request.POST.get('conselho_classe_investigador')


	# CODIGO NUMERO UNICO #
	
	date = datetime.now()
	dia = date.day
	mes = date.month
	ano = date.year

	dia = str(dia)
	mes = str(mes)
	ano = str(ano)

	dia = dia.zfill(2)
	mes = mes.zfill(2)
	ano = ano.zfill(4)

	date = dia + mes + ano
	
	ultimo_registro_caso = CasoEsporotricose.objects.all().last()
	if ultimo_registro_caso:
		if ultimo_registro_caso.numero_unico == "" or ultimo_registro_caso.numero_unico == None:
			codigo = "000000000000"
		else:
			codigo = ultimo_registro_caso.numero_unico
			print("codigo:", codigo)
	else:
		codigo = "000000000000"
	ano_codigo_anterior = codigo[4]+codigo[5]+codigo[6]+codigo[7] # Pegando o ano do codigo vindo do banco.
	
	
	if ano_codigo_anterior != ano: # Se o ano do codigo do banco for diferente do ano atual, zera o codigo.
		codigo = date + "0000"

	n_codigo = codigo[8]+codigo[9]+codigo[10]+codigo[11] # Pegando o codigo final de 4 digitos.
	n_codigo = int(n_codigo)
	n_codigo += 1
	n_codigo = str(n_codigo)
	n_codigo = n_codigo.zfill(4)
	codigo = date + n_codigo
		
	# CODIGO NUMERO UNICO #

	
	CasoEsporotricose.objects.create(
		responsavel_pelas_informacoes = responsavel_pelas_informacoes,
		responsavel_gerencia_operacional = responsavel_gerencia_operacional,
		responsavel_nucleo = responsavel_nucleo,
		responsavel_area_tecnica = responsavel_area_tecnica,
		responsavel_gerencia_regional = responsavel_gerencia_regional,
		responsavel_municipio = str(responsavel_municipio).upper(),
		tipo_notificacao = tipo_notificacao,
		agravo_doenca = agravo_doenca,
		codigo_cib10 = codigo_cib10,
		data_notificacao = data_notificacao,
		estado = estado,
		municipio = municipio,
		gerencia = gerencia,
		codigo_ibge = codigo_ibge,
		data_primeiros_sintomas = data_primeiros_sintomas,
		unidade_saude = unidade_saude,
		unidade_saude_outro = unidade_saude_outro,
		nome_paciente = nome_paciente,
		data_nascimento_paciente = data_nascimento_paciente,
		sexo_paciente = sexo_paciente,
		idade_paciente = idade_paciente,
		paciente_gestante = paciente_gestante,
		raca_paciente = raca_paciente,
		escolaridade_paciente = escolaridade_paciente,
		cartao_sus_paciente = cartao_sus_paciente,
		nome_mae_paciente = nome_mae_paciente,
		cep_residencia = cep_residencia,
		uf_residencia = uf_residencia,
		municipio_residencia = municipio_residencia,
		bairro_residencia = bairro_residencia,
		codigo_ibge_residencia = codigo_ibge_residencia,
		rua_residencia = rua_residencia,
		numero_residencia = numero_residencia,
		complemento_residencia = complemento_residencia,
		distrito_residencia = distrito_residencia,
		ponto_referencia_residencia = ponto_referencia_residencia,
		telefone_residencia = telefone_residencia,
		zona_residencia = zona_residencia,
		pais_residencia = pais_residencia,
		data_investigacao = data_investigacao,
		ocupacao = ocupacao,
		ambientes_frequentados = ambientes_frequentados,
		ambientes_frequentados_outros = ambientes_frequentados_outros,
		animais_que_teve_contato = animais_que_teve_contato,
		animais_que_teve_contato_outros = animais_que_teve_contato_outros,
		natureza_contato_animais = natureza_contato_animais,
		natureza_contato_animais_outros = natureza_contato_animais_outros,
		relacao_animal_doente = relacao_animal_doente,
		relacao_animal_doente_outros = relacao_animal_doente_outros,
		exerce_atividade_contato_plantas = exerce_atividade_contato_plantas,
		historico_contato_material = historico_contato_material,
		presenca_lesao_pele = presenca_lesao_pele,
		natureza_lesao = natureza_lesao,
		natureza_lesao_outro = natureza_lesao_outro,
		local_lesao = local_lesao,
		local_lesao_outro = local_lesao_outro,
		diagnostico_forma_extrac_doenca = diagnostico_forma_extrac_doenca,
		localizacao_forma_extrac_doenca = localizacao_forma_extrac_doenca,
		houve_coleta_material = houve_coleta_material,
		data_coleta1 = data_coleta1,
		numero_gal1 = numero_gal1,
		data_coleta2 = data_coleta2,
		numero_gal2 = numero_gal2,
		data_coleta3 = data_coleta3,
		numero_gal3 = numero_gal3,
		resultado_isolamento = resultado_isolamento,
		agente = agente,
		histopatologia = histopatologia,
		data_resultado_exame1 = data_resultado_exame1,
		descricao_exame_1 = descricao_exame_1,
		resultado_exame1 = resultado_exame1,
		data_resultado_exame2 = data_resultado_exame2,
		descricao_exame_2 = descricao_exame_2,
		resultado_exame2 = resultado_exame2,
		data_resultado_exame3 = data_resultado_exame3,
		descricao_exame_3 = descricao_exame_3,
		resultado_exame3 = resultado_exame3,
		data_inicio_tratamento1 = data_inicio_tratamento1,
		droga_administrada1 = droga_administrada1,
		esquema_terapeutico1 = esquema_terapeutico1,
		data_inicio_tratamento2 = data_inicio_tratamento2,
		droga_administrada2 = droga_administrada2,
		esquema_terapeutico2 = esquema_terapeutico2,
		data_inicio_tratamento3 = data_inicio_tratamento3,
		droga_administrada3 = droga_administrada3,
		esquema_terapeutico3 = esquema_terapeutico3,
		hospitalizacao = hospitalizacao,
		data_internacao = data_internacao,
		data_da_alta = data_da_alta,
		uf_hospitalizacao = uf_hospitalizacao,
		municipio_hospitalizacao = municipio_hospitalizacao,
		codigo_ibge_hospitalizacao = codigo_ibge_hospitalizacao,
		nome_hospital_hospitalizacao = nome_hospital_hospitalizacao,
		classificacao_final = classificacao_final,
		criterio_confirmacao = criterio_confirmacao,
		caso_autoctone_municipio_residencia = caso_autoctone_municipio_residencia,
		uf_caso_autoctone = uf_caso_autoctone,
		pais_caso_autoctone = pais_caso_autoctone,
		municipio_caso_autoctone = municipio_caso_autoctone,
		codigo_ibge_caso_autoctone = codigo_ibge_caso_autoctone,
		distrito_caso_autoctone = distrito_caso_autoctone,
		bairro_caso_autoctone = bairro_caso_autoctone,
		area_provavel_infeccao_caso_autoctone = area_provavel_infeccao_caso_autoctone,
		ambiente_infeccao_caso_autoctone = ambiente_infeccao_caso_autoctone,
		doenca_rel_trabalho_caso_autoctone = doenca_rel_trabalho_caso_autoctone,
		evolucao_caso = evolucao_caso,
		status_caso = status_caso,
		data_obito = data_obito,
		data_encerramento = data_encerramento,
		observacao = observacao,
		nome_investigador = nome_investigador,
		funcao_investigador = funcao_investigador,
		email_investigador = email_investigador,
		telefone_investigador = telefone_investigador,
		conselho_classe_investigador = conselho_classe_investigador,
		numero_unico = codigo

		)
	
	
	return redirect("/my_datas", messages = messages.success(request, 'Caso criado com sucesso!'))


# INDEX ABERTO

def index_aberto(request):
	return render(request, 'index_aberto.html')

def ajax_index_aberto(request):
	dados = CasoEsporotricose.objects.all()

	# Separando os dados em cinco colunas e pegando a quantidade de casos de cada um.
	detectados = len(dados.filter(Q(resultado_isolamento = 'DETECTADO') | Q(resultado_isolamento = 'DETECTÁVEL')))
	nao_detectados = len(dados.filter(Q(resultado_isolamento = 'NÃO DETECTADO') | Q(resultado_isolamento = 'NÃO DETECTÁVEL')))
	inconclusivo = len(dados.filter(resultado_isolamento = 'INCONCLUSIVO'))
	nao_realizado = len(dados.filter(resultado_isolamento = 'NÃO REALIZADO'))
	vazio = len(dados.filter(resultado_isolamento = None))
	
	# Total de casos.
	total_casos = "Total de Casos: " + str(detectados + nao_detectados + inconclusivo + nao_realizado + vazio)

	# Devolvendo os dados para o front.
	data = {
		'doenca':'Casos notificacos para Esporotricose Humana, segundo classificação final.',
		'detectados':detectados,
		'nao_detectados':nao_detectados,
		'inconclusivo':inconclusivo,
		'nao_realizado':nao_realizado,
		'vazio':vazio,
		'total':total_casos
	}
	return JsonResponse(data)





def ajax_filtrar_index_aberto(request):
	# Dicionário para receber o Value do agravo e retornar o nome de forma mais adequada.
	agravos = {'Selecione':'', 'ESPOROTRICOSE':'Casos notificacos para Esporotricose Humana, segundo classificação final.', 'ZIKA':'Zika', 'CHIKUNGUNYA':'Chikungunya', 'DENGUE':'Dengue'}

	# Capturando as informações
	ano = request.GET.get('ano')
	inicio = request.GET.get('inicio')
	fim = request.GET.get('fim')
	agravo = request.GET.get('agravo')
	
	if agravo != 'Selecione':

		# Se o agravo for diferente de 'Selecione' filtra os dados pelo agravo.
		dados = CasoEsporotricose.objects.all().filter(agravo_doenca = agravo)
		
		# Filtrando por componentes preenchidos.
		if inicio == "" and fim == "" and ano != "":
			dados = dados.filter(data_notificacao__year=ano)
		# Filtrando por componentes preenchidos.
		elif inicio != "" and fim != "":
			dados = dados.filter(data_notificacao__range=[inicio,fim])
		# Filtrando por componentes preenchidos.
		elif inicio != "" and fim == "":
			dados = dados.filter(data_notificacao=inicio)
		# Filtrando por componentes preenchidos.
		elif inicio == "" and fim != "":
			dados = dados.filter(data_notificacao=fim)

		# Separando os dados em cinco colunas e pegando a quantidade de casos de cada um.
		detectados = len(dados.filter(Q(resultado_isolamento = 'DETECTADO') | Q(resultado_isolamento = 'DETECTÁVEL')))
		nao_detectados = len(dados.filter(Q(resultado_isolamento = 'NÃO DETECTADO') | Q(resultado_isolamento = 'NÃO DETECTÁVEL')))
		inconclusivo = len(dados.filter(resultado_isolamento = 'INCONCLUSIVO'))
		nao_realizado = len(dados.filter(resultado_isolamento = 'NÃO REALIZADO'))
		vazio = len(dados.filter(resultado_isolamento = None))
		
		# Total de casos.
		total_casos = "Total de Casos: " + str(detectados + nao_detectados + inconclusivo + nao_realizado + vazio)

		# Devolvendo os dados para o front.
		data = {
			'doenca':agravos[agravo],
			'detectados':detectados,
			'nao_detectados':nao_detectados,
			'inconclusivo':inconclusivo,
			'nao_realizado':nao_realizado,
			'vazio':vazio,
			'total':total_casos
		}

		return JsonResponse(data)
	# Caso o agravo não esteja selecionado.
	else:
		data = {
			'doenca':'',
			'detectados':0,
			'nao_detectados':0,
			'inconclusivo':0,
			'nao_realizado':0,
			'vazio':0,
			'total':0
		}
		return JsonResponse(data)
	
def ajax_exportar_index_fechado(request):
	ano = request.POST.get('ano')
	inicio = request.POST.get('inicio')
	fim = request.POST.get('fim')
	option = request.POST.get('export_select')

	
	if inicio == "" and fim == "" and ano == "":
		dados = CasoEsporotricose.objects.all()
		
	if inicio == "" and fim == "" and ano != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__year=ano)
		
	if inicio != "" and fim == "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[inicio,inicio])
		
	if inicio == "" and fim != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[fim,fim])
		
	if inicio != "" and fim != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[inicio,fim])
		
	data = []
	for item in dados:
		data.append([item.tipo_notificacao, item.agravo_doenca, item.codigo_cib10, item.data_notificacao, item.estado, item.municipio, item.codigo_ibge, item.data_primeiros_sintomas, item.unidade_saude, item.nome_paciente, item.data_nascimento_paciente, item.idade_paciente, item.sexo_paciente, item.paciente_gestante, item.raca_paciente, item.escolaridade_paciente, item.cartao_sus_paciente, item.nome_mae_paciente, item.cep_residencia, item.uf_residencia, item.municipio_residencia, item.bairro_residencia, item.codigo_ibge_residencia, item.rua_residencia, item.numero_residencia, item.complemento_residencia, item.distrito_residencia, item.ponto_referencia_residencia, item.telefone_residencia, item.zona_residencia, item.pais_residencia, item.data_investigacao, item.ocupacao, item.ambientes_frequentados, item.animais_que_teve_contato, item.natureza_contato_animais, item.relacao_animal_doente, item.exerce_atividade_contato_plantas, item.historico_contato_material, item.presenca_lesao_pele, item.natureza_lesao, item.local_lesao, item.diagnostico_forma_extrac_doenca, item.localizacao_forma_extrac_doenca, item.houve_coleta_material, item.data_coleta1, item.numero_gal1, item.data_coleta2, item.numero_gal2, item.data_coleta3, item.numero_gal3, item.resultado_isolamento, item.agente, item.histopatologia, item.data_resultado_exame1, item.descricao_exame_1, item.resultado_exame1, item.data_resultado_exame2, item.descricao_exame_2, item.resultado_exame2, item.data_resultado_exame3, item.descricao_exame_3, item.resultado_exame3, item.data_inicio_tratamento1, item.droga_administrada1, item.esquema_terapeutico1, item.data_inicio_tratamento2, item.droga_administrada2, item.esquema_terapeutico2, item.data_inicio_tratamento3, item.droga_administrada3, item.esquema_terapeutico3, item.hospitalizacao, item.data_internacao, item.data_da_alta, item.uf_hospitalizacao, item.municipio_hospitalizacao, item.codigo_ibge_hospitalizacao, item.nome_hospital_hospitalizacao, item.classificacao_final, item.criterio_confirmacao, item.caso_autoctone_municipio_residencia, item.uf_caso_autoctone, item.pais_caso_autoctone, item.municipio_caso_autoctone, item.codigo_ibge_caso_autoctone, item.distrito_caso_autoctone, item.bairro_caso_autoctone, item.area_provavel_infeccao_caso_autoctone, item.ambiente_infeccao_caso_autoctone, item.doenca_rel_trabalho_caso_autoctone, item.evolucao_caso, item.data_obito, item.data_encerramento, item.observacao, item.nome_investigador, item.funcao_investigador, item.email_investigador, item.telefone_investigador, item.conselho_classe_investigador, item.responsavel_pelas_informacoes_id, item.unidade_saude_outro, item.gerencia_id])
	
	data = pd.DataFrame(data, columns=['Tipo de Notificação', 'Agravo/Doença', 'Codigo CID-10', 'Data de Notificação', 'Estado', 'Municipio', 'Codigo IBGE', 'Data dos Primeiros Sintomas', 'Unidade de Saúde', 'Nome do Paciente', 'Data de Nascimento do Paciente', 'Idade do Paciente', 'Sexo do Paciente', 'Paciente Gestante', 'Raça do Paciente', 'Escolaridade do Paciente', 'Cartão SUS do Paciente', 'Nome da Mãe do Paciente', 'CEP da Residência', 'UF da Residência', 'Municipio da Residência', 'Bairro da Residência', 'Codigo IBGE da Residência', 'Rua da Residência', 'Numero_da Residência', 'Complemento da Residência', 'Distrito da Residência', 'Ponto de Referência da Residência', 'Telefone da Residência', 'Zona da Residência', 'País da Residência', 'Data de Investigação', 'Ocupação', 'Ambientes Frequentados', 'Animais que Teve Contato', 'Natureza do Contato com os Animais', 'Relação com o Animal Doente', 'Exerce Atividade Contato Plantas', 'Histórico Contato Material', 'Presença de Lesão na Pele', 'Natureza da Lesão', 'Local da Lesão', 'Diagnóstico Forma Extrac Doença', 'Localização Forma Extrac Doença', 'Houve Coleta de Material', 'Data da Coleta 1', 'Numero Gal 1', 'Data da Coleta 2', 'Numero Gal 2', 'Data da Coleta 3', 'Numero Gal 3', 'Resultado de Isolamento', 'Agente', 'Histopatologia', 'Data do Restultado EXAME 1', 'Descrição EXAME 1', 'Resultado EXAME 1', 'Data do Resultado EXAME 2', 'Descrição EXAME 2', 'Resultado EXAME 2', 'Data do Resultado EXAME 3', 'Descrição EXAME 3', 'Resultado EXAME 3', 'Data de Início do Tratamento 1', 'Droga Administrada 1', 'Esquema Terapeutico 1', 'Data de Início do Tratamento 2', 'Droga Administrada 2', 'Esquema Terapeutico 2', 'Data de Início do Tratamento 3', 'Droga Administrada 3', 'Esquema Terapeutico 3', 'Hospitalizacão', 'Data de Internação', 'Data da Alta', 'UF da Hospitalização', 'Município da Hospitalização', 'Codigo IBGE da Hospitalização', 'Nome do Hospital da Hospitalização', 'Classificação Final', 'Critério de Confirmação', 'Caso Autoctone Municipio Residência', 'UF Caso Autoctone', 'País Caso Autoctone', 'Município Caso Autoctone', 'Codigo IBGE Caso Autoctone', 'Distrito Caso Autoctone', 'Bairro Caso Autoctone', 'Área Provável de Infecção Caso Autoctone', 'Ambiente Infecção Caso Autoctone', 'Doença Rel Trabalho Caso Autoctone', 'Evolução do Caso', 'Data do Óbito', 'Data de Encerramento', 'Observação', 'Nome do Investigador', 'Função do Investigador', 'Email do Investigador', 'Telefone do Investigador', 'Conselho Classe Investigador', 'Responsável Pelas Informações ID', 'Unidade de Saude Outro', 'Gerencia ID'])
	
	if option == "csv":
		response = HttpResponse(content_type = "text/csv")
		response['Content-Disposition'] = 'attachment; filename=casos_esporotricose.csv'
		data.to_csv(response, index=False)

		return response

	if option =="excel":
		response = HttpResponse(content_type = "application/ms-excel")
		response['Content-Disposition'] = 'attachment; filename=casos_esporotricose.xlsx'

		data.to_excel(response, index=False)
		return response
	

def ajax_exportar_index_aberto(request):
	ano = request.POST.get('ano')
	inicio = request.POST.get('inicio')
	fim = request.POST.get('fim')
	option = request.POST.get('export_select')
	
	if inicio == "" and fim == "" and ano == "":
		dados = CasoEsporotricose.objects.all()
		
	if inicio == "" and fim == "" and ano != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__year=ano)
		
	if inicio != "" and fim == "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[inicio,inicio])
		
	if inicio == "" and fim != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[fim,fim])
		
	if inicio != "" and fim != "":
		dados = CasoEsporotricose.objects.filter(data_notificacao__range=[inicio,fim])
		
	data = []
	for item in dados:
		data.append([item.tipo_notificacao, item.agravo_doenca, item.codigo_cib10, item.data_notificacao, item.estado, item.codigo_ibge, item.data_primeiros_sintomas, item.unidade_saude, item.idade_paciente, item.sexo_paciente, item.paciente_gestante, item.raca_paciente, item.escolaridade_paciente,  item.uf_residencia, item.municipio_residencia,  item.codigo_ibge_residencia, item.distrito_residencia, item.zona_residencia, item.pais_residencia, item.data_investigacao, item.ocupacao, item.ambientes_frequentados, item.animais_que_teve_contato, item.natureza_contato_animais, item.relacao_animal_doente, item.exerce_atividade_contato_plantas, item.historico_contato_material, item.presenca_lesao_pele, item.natureza_lesao, item.local_lesao, item.diagnostico_forma_extrac_doenca, item.localizacao_forma_extrac_doenca, item.houve_coleta_material, item.data_coleta1, item.data_coleta2, item.data_coleta3, item.resultado_isolamento, item.agente, item.histopatologia, item.data_resultado_exame1, item.descricao_exame_1, item.resultado_exame1, item.data_resultado_exame2, item.descricao_exame_2, item.resultado_exame2, item.data_resultado_exame3, item.descricao_exame_3, item.resultado_exame3, item.data_inicio_tratamento1, item.droga_administrada1, item.esquema_terapeutico1, item.data_inicio_tratamento2, item.droga_administrada2, item.esquema_terapeutico2, item.data_inicio_tratamento3, item.droga_administrada3, item.esquema_terapeutico3, item.hospitalizacao, item.data_internacao, item.data_da_alta, item.uf_hospitalizacao, item.municipio_hospitalizacao, item.codigo_ibge_hospitalizacao, item.nome_hospital_hospitalizacao, item.classificacao_final, item.criterio_confirmacao, item.caso_autoctone_municipio_residencia, item.uf_caso_autoctone, item.pais_caso_autoctone, item.municipio_caso_autoctone, item.codigo_ibge_caso_autoctone, item.distrito_caso_autoctone, item.bairro_caso_autoctone, item.area_provavel_infeccao_caso_autoctone, item.ambiente_infeccao_caso_autoctone, item.doenca_rel_trabalho_caso_autoctone, item.evolucao_caso, item.data_obito, item.data_encerramento, item.observacao, item.unidade_saude_outro])
	
	data = pd.DataFrame(data, columns=['Tipo de Notificação', 'Agravo/Doença', 'Codigo CID-10', 'Data de Notificação', 'Estado', 'Codigo IBGE', 'Data dos Primeiros Sintomas', 'Unidade de Saúde', 'Idade do Paciente', 'Sexo do Paciente', 'Paciente Gestante', 'Raça do Paciente', 'Escolaridade do Paciente', 'UF da Residência', 'Municipio da Residência',  'Codigo IBGE da Residência', 'Distrito da Residência', 'Zona da Residência', 'País da Residência', 'Data de Investigação', 'Ocupação', 'Ambientes Frequentados', 'Animais que Teve Contato', 'Natureza do Contato com os Animais', 'Relação com o Animal Doente', 'Exerce Atividade Contato Plantas', 'Histórico Contato Material', 'Presença de Lesão na Pele', 'Natureza da Lesão', 'Local da Lesão', 'Diagnóstico Forma Extrac Doença', 'Localização Forma Extrac Doença', 'Houve Coleta de Material', 'Data da Coleta 1', 'Data da Coleta 2', 'Data da Coleta 3', 'Resultado de Isolamento', 'Agente', 'Histopatologia', 'Data do Restultado EXAME 1', 'Descrição EXAME 1', 'Resultado EXAME 1', 'Data do Resultado EXAME 2', 'Descrição EXAME 2', 'Resultado EXAME 2', 'Data do Resultado EXAME 3', 'Descrição EXAME 3', 'Resultado EXAME 3', 'Data de Início do Tratamento 1', 'Droga Administrada 1', 'Esquema Terapeutico 1', 'Data de Início do Tratamento 2', 'Droga Administrada 2', 'Esquema Terapeutico 2', 'Data de Início do Tratamento 3', 'Droga Administrada 3', 'Esquema Terapeutico 3', 'Hospitalizacão', 'Data de Internação', 'Data da Alta', 'UF da Hospitalização', 'Município da Hospitalização', 'Codigo IBGE da Hospitalização', 'Nome do Hospital da Hospitalização', 'Classificação Final', 'Critério de Confirmação', 'Caso Autoctone Municipio Residência', 'UF Caso Autoctone', 'País Caso Autoctone', 'Município Caso Autoctone', 'Codigo IBGE Caso Autoctone', 'Distrito Caso Autoctone', 'Bairro Caso Autoctone', 'Área Provável de Infecção Caso Autoctone', 'Ambiente Infecção Caso Autoctone', 'Doença Rel Trabalho Caso Autoctone', 'Evolução do Caso', 'Data do Óbito', 'Data de Encerramento', 'Observação', 'Unidade de Saude Outro'])
	
	if option == "csv":
		response = HttpResponse(content_type = "text/csv")
		response['Content-Disposition'] = 'attachment; filename=casos_esporotricose_humana.csv'
		data.to_csv(response, index=False)

		return response

	elif option =="excel":
		response = HttpResponse(content_type = "application/ms-excel")
		response['Content-Disposition'] = 'attachment; filename=casos_esporotricose_humana.xlsx'

		data.to_excel(response, index=False)
		return response
	else:
		return redirect('index')

# INDEX ABERTO

@login_required(login_url='/login/')
def caso_esporotricose_edit(request, id):
	caso = CasoEsporotricose.objects.get(id=id)
	estados = Estado.objects.all().order_by('nome')
	print('UF = ', caso.uf_residencia)
	try:
		estado_caso_str = int(caso.uf_residencia)
	except Exception as e:
		estado_caso = None
	else:
		estado_caso = Estado.objects.get(id=estado_caso_str)
			
	cidade_caso_ = caso.municipio_residencia
	
	print("municipio do caso:", cidade_caso_, type(cidade_caso_))

	cidade_caso_registro = Municipio.objects.filter(nome=cidade_caso_)

	print("cidade_caso_registro:", cidade_caso_registro)

	municipio_residencia_br = Municipios.objects.filter(nome=cidade_caso_)


	#cidade_caso = int(caso.municipio_residencia)
	cidade_caso = municipio_residencia_br
	codigo_ibge_residencia = caso.codigo_ibge_residencia
	
	if caso.municipio_residencia != None or caso.municipio_residencia != '':
		#municipio_residencia = (caso.municipio_residencia).title()
		municipio_residencia = caso.municipio_residencia
	else:
		municipio_residencia = None
	
	
	cidade_caso_id = cidade_caso

	#cidade_caso = Municipios.objects.get(id=cidade_caso_id)

	print("residencia do caso:", estado_caso)

	municipios = Municipio.objects.all().order_by('nome')
	unidades_saude = []
	codigos_ibge = []
	#print(caso.codigo_ibge_caso_autoctone)

	if caso.municipio:
		municipio_caso = caso.municipio
		codigo_ibge = CodigoIbge.objects.get(municipio=municipio_caso)
		unidade_saude_caso = UnidadeSaude.objects.filter(municipio=municipio_caso)
		unidade_saude_caso = caso.unidade_saude
	else:
		codigo_ibge = None
		unidade_saude_caso = None

	
	if caso.data_notificacao != None:
		caso.data_notificacao = datetime.strftime(caso.data_notificacao, '%Y-%m-%d')
	else:
		caso.data_notificacao = ""
	
	if caso.data_primeiros_sintomas != None:
		caso.data_primeiros_sintomas = datetime.strftime(caso.data_primeiros_sintomas, '%Y-%m-%d')
	else:
		caso.data_primeiros_sintomas = ""
	
	if caso.data_nascimento_paciente != None:
		caso.data_nascimento_paciente = datetime.strftime(caso.data_nascimento_paciente, '%Y-%m-%d')
	else:
		caso.data_nascimento_paciente = ""

	if caso.data_resultado_exame1 != None:
		caso.data_resultado_exame1 = datetime.strftime(caso.data_resultado_exame1, '%Y-%m-%d')
	else:
		caso.data_resultado_exame1 = ""

	if caso.data_resultado_exame2 != None:
		caso.data_resultado_exame2 = datetime.strftime(caso.data_resultado_exame2, '%Y-%m-%d')
	else:
		caso.data_resultado_exame2 = ""

	if caso.data_resultado_exame3 != None:
		caso.data_resultado_exame3 = datetime.strftime(caso.data_resultado_exame3, '%Y-%m-%d')
	else:
		caso.data_resultado_exame3 = ""

	if caso.data_coleta1 != None:
		caso.data_coleta1 = datetime.strftime(caso.data_coleta1, '%Y-%m-%d')
	else:
		caso.data_coleta1 = ""

	if caso.data_coleta2 != None:
		caso.data_coleta2 = datetime.strftime(caso.data_coleta2, '%Y-%m-%d')
	else:
		caso.data_coleta2 = ""

	if caso.data_coleta3 != None:
		caso.data_coleta3 = datetime.strftime(caso.data_coleta3, '%Y-%m-%d')
	else:
		caso.data_coleta3 = ""

	if caso.data_investigacao != None:
		caso.data_investigacao = datetime.strftime(caso.data_investigacao, '%Y-%m-%d')
	else:
		caso.data_investigacao = ""

	if caso.data_inicio_tratamento1 != None:
		caso.data_inicio_tratamento1 = datetime.strftime(caso.data_inicio_tratamento1, '%Y-%m-%d')
	else:
		caso.data_inicio_tratamento1 = ""

	if caso.data_inicio_tratamento2 != None:
		caso.data_inicio_tratamento2 = datetime.strftime(caso.data_inicio_tratamento2, '%Y-%m-%d')
	else:
		caso.data_inicio_tratamento2 = ""

	if caso.data_inicio_tratamento3 != None:
		caso.data_inicio_tratamento3 = datetime.strftime(caso.data_inicio_tratamento3, '%Y-%m-%d')
	else:
		caso.data_inicio_tratamento3 = ""

	if caso.data_internacao != None:
		caso.data_internacao = datetime.strftime(caso.data_internacao, '%Y-%m-%d')
	else:
		caso.data_internacao = ""

	if caso.data_da_alta != None:
		caso.data_da_alta = datetime.strftime(caso.data_da_alta, '%Y-%m-%d')
	else:
		caso.data_da_alta = ""

	if caso.data_obito != None:
		caso.data_obito = datetime.strftime(caso.data_obito, '%Y-%m-%d')
	else:
		caso.data_obito = ""

	if caso.data_encerramento != None:
		caso.data_encerramento = datetime.strftime(caso.data_encerramento, '%Y-%m-%d')
	else:
		caso.data_encerramento = ""
		
	return render(request, 'caso_esporotricose_edit.html', {'form':caso, 'municipios':municipios, 'unidades_saude':unidades_saude, 
		'codigos_ibge':codigos_ibge, 'estados':estados, 'codigo_ibge':codigo_ibge, 'unidade_saude_caso':unidade_saude_caso,
		'estado_caso':estado_caso, 'cidade_caso':cidade_caso, 'codigo_ibge_residencia': codigo_ibge_residencia, 
		'municipio_residencia': municipio_residencia})

@login_required(login_url='/login/')
def set_caso_esporotricose_edit(request, id):
	# Dados do responsável pela edição do caso.
	responsavel_pelas_informacoes = request.user.id
	responsavel_gerencia_operacional = request.user.gerencia_operacional
	responsavel_nucleo = request.user.nucleo
	responsavel_area_tecnica = request.user.area_tecnica
	responsavel_gerencia_regional = request.user.area_tecnica
	responsavel_municipio = request.user.municipio.nome

	#Dados Gerais
	tipo_notificacao = request.POST.get('tipo_notificacao')
	agravo_doenca = request.POST.get('agravo_doenca')
	codigo_cib10 = request.POST.get('codigo_cid')
	
	data_notificacao_cap = request.POST.get('data_notificacao')
	
	if data_notificacao_cap == '' or data_notificacao_cap == None:
		data_notificacao = None
	else:
		data_notificacao = datetime.strptime(data_notificacao_cap, '%Y-%m-%d').date()

	
	estado = request.POST.get('uf_notificacao')
	
	municipio_cap = request.POST.get('municipio_notificacao')
	if municipio_cap == '' or municipio_cap == None:
		municipio = None
	else:
		municipio = int(municipio_cap)

	municipio_id = municipio
	municipio_ = Municipio.objects.get(id=municipio_id)

	gerencia = municipio_.gerencia

	codigo_ibge = request.POST.get('codigo_ibge_dados_gerais')
	
	data_primeiros_sintomas_cap = request.POST.get('data_primeiros_sintomas')
	if data_primeiros_sintomas_cap == '' or data_primeiros_sintomas_cap == None:
		data_primeiros_sintomas = None
	else:
		data_primeiros_sintomas = datetime.strptime(data_primeiros_sintomas_cap, '%Y-%m-%d').date()

	
	unidade_saude = request.POST.get('unidade_saude')
	unidade_saude_outro = request.POST.get('unidade_saude_outro')
	
	#notificação individual
	nome_paciente = request.POST.get('nome_paciente').upper()
	
	data_nascimento_paciente_cap = request.POST.get('data_nasc')
	if data_nascimento_paciente_cap == '' or data_nascimento_paciente_cap == None:
		data_nascimento_paciente = None
	else:
		data_nascimento_paciente = datetime.strptime(data_nascimento_paciente_cap, '%Y-%m-%d').date()

	sexo_paciente = request.POST.get('sexo')
	
	idade_paciente_cap = request.POST.get('result')
	if idade_paciente_cap == '' or idade_paciente_cap == None or idade_paciente_cap == "None":
		idade_paciente = None
	else:
		idade_paciente = int(idade_paciente_cap)

	paciente_gestante = request.POST.get('gestacao')
	raca_paciente = request.POST.get('raca')
	escolaridade_paciente = request.POST.get('escolaridade')
	cartao_sus_paciente = request.POST.get('cartao_sus')
	nome_mae_paciente = request.POST.get('nome_mae').upper()
	
	#Dados Residencia
	cep_residencia = request.POST.get('cep_residencia')
	uf_residencia = request.POST.get('uf_residencia')
	print('uf_residencia: ', uf_residencia)	
	municipio_residencia = request.POST.get('cidade_residencia')
	
	if uf_residencia == '12': # Se o uf for PB pega dados no modelo Municipio
		try:
			municipio_residencia = Municipio.objects.get(id=municipio_residencia)
			municipio_residencia = municipio_residencia.nome
		except:
			pass
	else: # Caso contrário pega os dados no modelo Municipios
		try:
			municipio_residencia = Municipios.objects.get(id=municipio_residencia)
			municipio_residencia = municipio_residencia.nome
		except:
			pass	
	municipio_residencia = municipio_residencia
	
	bairro_residencia = request.POST.get('bairro_residencia')
	codigo_ibge_residencia = request.POST.get('codigo_ibge_residencia')
	rua_residencia = request.POST.get('rua_residencia')
	numero_residencia = request.POST.get('numero_residencia')
	complemento_residencia = request.POST.get('complemento_residencia')
	distrito_residencia = request.POST.get('distrito_residencia')
	ponto_referencia_residencia = request.POST.get('ponto_referencia_residencia')
	telefone_residencia = request.POST.get('telefone_residencia')
	zona_residencia = request.POST.get('zona_residencia')
	pais_residencia = request.POST.get('pais_residente_fora_pais')
	
	#Antecedentes Epidemiologicos
	data_investigacao_cap = request.POST.get('data_investigacao')
	if data_investigacao_cap == '' or data_investigacao_cap == None:
		data_investigacao = None
	else:
		data_investigacao = datetime.strptime(data_investigacao_cap, '%Y-%m-%d').date()


	ocupacao = request.POST.get('ocupacao')
	ambientes_frequentados = request.POST.getlist('ambientes_frequentados')
	animais_que_teve_contato = request.POST.getlist('animais_que_teve_contato')
	natureza_contato_animais = request.POST.getlist('natureza_contato_animais')
	relacao_animal_doente = request.POST.getlist('relacao_animal_doente')
	exerce_atividade_contato_plantas = request.POST.get('exerc_ativ_contato_plantas')
	historico_contato_material = request.POST.get('lesao_manuseio')
	
	#dados clínicos
	presenca_lesao_pele = request.POST.get('presenca_lesao_pele')
	natureza_lesao = request.POST.getlist('natureza_lesao')
	natureza_lesao_outro = request.POST.get('natureza_lesao_outros')
	local_lesao = request.POST.getlist('local_lesao')
	local_lesao_outro = request.POST.get('local_lesao_outros')
	diagnostico_forma_extrac_doenca = request.POST.get('diagnostico_extracutaneo')
	localizacao_forma_extrac_doenca = request.POST.get('local_extra_cutanea')
	
	#Dados Laboratoriais
	houve_coleta_material = request.POST.get('coleta_exame_labo_sim')
	
	data_coleta1 = request.POST.get('data_coleta_mat_exame1')
	if data_coleta1 == '' or data_coleta1 == None:
		data_coleta1 = None
		numero_gal1 = None
	else:
		data_coleta1 = datetime.strptime(data_coleta1, '%Y-%m-%d').date()
		numero_gal1 = request.POST.get('numero_gal1')

	data_coleta2 = request.POST.get('data_coleta_mat_exame2')
	if data_coleta2 == '' or data_coleta2 == None:
		data_coleta2 = None
		numero_gal2 = None
	else:
		data_coleta2 = datetime.strptime(data_coleta2, '%Y-%m-%d').date()
		numero_gal2 = request.POST.get('numero_gal2')

	data_coleta3 = request.POST.get('data_coleta_mat_exame3')
	if data_coleta3 == '' or data_coleta3 == None:
		data_coleta3 = None
		numero_gal3 = None
	else:
		data_coleta3 = datetime.strptime(data_coleta3, '%Y-%m-%d').date()
		numero_gal3 = request.POST.get('numero_gal3')
	
	resultado_isolamento = request.POST.get('resultado_isolamento')
	agente = request.POST.get('agente_isolamento')
	histopatologia = request.POST.get('histopatologia')
	
	data_exame1_cap = request.POST.get('data_outros_exames_1')
	if data_exame1_cap == '' or data_exame1_cap == None:
		data_resultado_exame1 = None
	else:
		data_resultado_exame1 = datetime.strptime(data_exame1_cap, '%Y-%m-%d').date()
	
	descricao_exame_1 = request.POST.get('descricao_outros_exames_1')
	resultado_exame1 = request.POST.get('resultado_outros_exames_1')
	
	data_exame2_cap = request.POST.get('data_outros_exames_2')
	if data_exame2_cap == '' or data_exame2_cap == None:
		data_resultado_exame2 = None
	else:
		data_resultado_exame2 = datetime.strptime(data_exame2_cap, '%Y-%m-%d').date()

	descricao_exame_2 = request.POST.get('descricao_outros_exames_2')
	resultado_exame2 = request.POST.get('resultado_outros_exames_2')
	
	
	data_exame3_cap = request.POST.get('data_outros_exames_3')
	if data_exame3_cap == '' or data_exame3_cap == None:
		data_resultado_exame3 = None
	else:
		data_resultado_exame3 = datetime.strptime(data_exame3_cap, '%Y-%m-%d').date()
		
	descricao_exame_3 = request.POST.get('descricao_outros_exames_3')
	resultado_exame3 = request.POST.get('resultado_outros_exames_3')
	
	
	#tratamento
	data_inicio_tratamento1_cap = request.POST.get('data_inicio_tratamento')
	if data_inicio_tratamento1_cap == '' or data_inicio_tratamento1_cap == None:
		data_inicio_tratamento1 = None		
	else:
		data_inicio_tratamento1 = datetime.strptime(data_inicio_tratamento1_cap, '%Y-%m-%d').date()
	

	droga_administrada1 = request.POST.get('droga_admin_tratamento')
	esquema_terapeutico1 = request.POST.get('esquema_terapeutico')
	
	data_inicio_tratamento2_cap = request.POST.get('data_inicio_tratamento2')
	if data_inicio_tratamento2_cap == '' or data_inicio_tratamento2_cap == None:
		data_inicio_tratamento2 = None
	else:
		data_inicio_tratamento2 = datetime.strptime(data_inicio_tratamento2_cap, '%Y-%m-%d').date()
	
	
	droga_administrada2 = request.POST.get('droga_admin_tratamento2')
	esquema_terapeutico2 = request.POST.get('esquema_terapeutico2')
	
	data_inicio_tratamento3_cap = request.POST.get('data_inicio_tratamento3')
	if data_inicio_tratamento3_cap == '' or data_inicio_tratamento3_cap == None:
		data_inicio_tratamento3 = None
	else:
		data_inicio_tratamento3 = datetime.strptime(data_inicio_tratamento3_cap, '%Y-%m-%d').date()
	

	droga_administrada3 = request.POST.get('droga_admin_tratamento3')
	esquema_terapeutico3 = request.POST.get('esquema_terapeutico3')
	hospitalizacao = request.POST.get('hospitalizado')
	
	data_internacao_cap = request.POST.get('data_internacao')
	if data_internacao_cap == '' or data_internacao_cap == None:
		data_internacao = None
	else:
		data_internacao = datetime.strptime(data_internacao_cap, '%Y-%m-%d').date()
	

	data_da_alta_cap = request.POST.get('data_alta')
	if data_da_alta_cap == '' or data_da_alta_cap == None:
		data_da_alta = None
	else:
		data_da_alta = datetime.strptime(data_da_alta_cap, '%Y-%m-%d').date()
	

	uf_hospitalizacao = request.POST.get('uf_internacao')
	if uf_hospitalizacao == 'PB':
		municipio_hospitalizacao = request.POST.get('municipio_hospital')
		codigo_ibge_hospitalizacao = request.POST.get('codigo_ibge_municipio_hospitalizacao')
		nome_hospital_hospitalizacao = request.POST.get('nome_hospital_hospitalizacao')
	else:
		municipio_hospitalizacao = request.POST.get('municipio_hospital_outro')
		codigo_ibge_hospitalizacao = request.POST.get('codigo_ibge_municipio_hospitalizacao_outro')
		nome_hospital_hospitalizacao = request.POST.get('nome_hospital_hospitalizacao_outro')
	
	#conclusao
	classificacao_final = request.POST.get('classificacao')
	criterio_confirmacao = request.POST.get('criterio')
	caso_autoctone_municipio_residencia = request.POST.get('autoctone')
	uf_caso_autoctone = request.POST.get('uf_caso_autoctone_municipio')
	pais_caso_autoctone = request.POST.get('pais_autoctone_municipio')
	municipio_caso_autoctone = request.POST.get('municipio_autoctone')
	codigo_ibge_caso_autoctone = request.POST.get('ibge_autoctone')
	distrito_caso_autoctone = request.POST.get('distrito_autoctone')
	bairro_caso_autoctone = request.POST.get('bairro_autoctone')
	area_provavel_infeccao_caso_autoctone = request.POST.get('area_provavel_infeccao')
	ambiente_infeccao_caso_autoctone = request.POST.get('ambiente_infeccao_autoctone')
	doenca_rel_trabalho_caso_autoctone = request.POST.get('doenca_rel_trabalho_autoctone')
	evolucao_caso = request.POST.get('evolucao_caso_autoctone')
	status_caso = request.POST.get('status_caso')
	
	data_obito_cap = request.POST.get('data_obito')
	if data_obito_cap == '' or data_obito_cap == None:
		data_obito = None
	else:
		data_obito = datetime.strptime(data_obito_cap, '%Y-%m-%d').date()
	

	data_encerramento_cap = request.POST.get('data_encerramento')
	if data_encerramento_cap == '' or data_encerramento_cap == None:
		data_encerramento = None		
	else:
		data_encerramento = datetime.strptime(data_encerramento_cap, '%Y-%m-%d').date()
	
	
	#observação
	observacao = request.POST.get('observacao')
	
	#investigador
	nome_investigador = request.POST.get('nome_investigador')
	funcao_investigador = request.POST.get('funcao_investigador')
	email_investigador = request.POST.get('email_investigador')
	telefone_investigador = request.POST.get('telefone_investigador')
	conselho_classe_investigador = request.POST.get('conselho_classe_investigador')

	CasoEsporotricose.objects.filter(id = id).update(
		responsavel_edicao = responsavel_pelas_informacoes,
		tipo_notificacao = tipo_notificacao,
		agravo_doenca = agravo_doenca,
		codigo_cib10 = codigo_cib10,
		data_notificacao = data_notificacao,
		estado = estado,
		municipio = municipio,
		gerencia = gerencia,
		codigo_ibge = codigo_ibge,
		data_primeiros_sintomas = data_primeiros_sintomas,
		unidade_saude = unidade_saude,
		unidade_saude_outro = unidade_saude_outro,
		nome_paciente = nome_paciente,
		data_nascimento_paciente = data_nascimento_paciente,
		sexo_paciente = sexo_paciente,
		idade_paciente = idade_paciente,
		paciente_gestante = paciente_gestante,
		raca_paciente = raca_paciente,
		escolaridade_paciente = escolaridade_paciente,
		cartao_sus_paciente = cartao_sus_paciente,
		nome_mae_paciente = nome_mae_paciente,
		cep_residencia = cep_residencia,
		uf_residencia = uf_residencia,
		municipio_residencia = municipio_residencia,
		bairro_residencia = bairro_residencia,
		codigo_ibge_residencia = codigo_ibge_residencia,
		rua_residencia = rua_residencia,
		numero_residencia = numero_residencia,
		complemento_residencia = complemento_residencia,
		distrito_residencia = distrito_residencia,
		ponto_referencia_residencia = ponto_referencia_residencia,
		telefone_residencia = telefone_residencia,
		zona_residencia = zona_residencia,
		pais_residencia = pais_residencia,
		data_investigacao = data_investigacao,
		ocupacao = ocupacao,
		ambientes_frequentados = ambientes_frequentados,
		animais_que_teve_contato = animais_que_teve_contato,
		natureza_contato_animais = natureza_contato_animais,
		relacao_animal_doente = relacao_animal_doente,
		exerce_atividade_contato_plantas = exerce_atividade_contato_plantas,
		historico_contato_material = historico_contato_material,
		presenca_lesao_pele = presenca_lesao_pele,
		natureza_lesao = natureza_lesao,
		natureza_lesao_outro = natureza_lesao_outro,
		local_lesao = local_lesao,
		local_lesao_outro = local_lesao_outro,
		diagnostico_forma_extrac_doenca = diagnostico_forma_extrac_doenca,
		localizacao_forma_extrac_doenca = localizacao_forma_extrac_doenca,
		houve_coleta_material = houve_coleta_material,
		data_coleta1 = data_coleta1,
		numero_gal1 = numero_gal1,
		data_coleta2 = data_coleta2,
		numero_gal2 = numero_gal2,
		data_coleta3 = data_coleta3,
		numero_gal3 = numero_gal3,
		resultado_isolamento = resultado_isolamento,
		agente = agente,
		histopatologia = histopatologia,
		data_resultado_exame1 = data_resultado_exame1,
		descricao_exame_1 = descricao_exame_1,
		resultado_exame1 = resultado_exame1,
		data_resultado_exame2 = data_resultado_exame2,
		descricao_exame_2 = descricao_exame_2,
		resultado_exame2 = resultado_exame2,
		data_resultado_exame3 = data_resultado_exame3,
		descricao_exame_3 = descricao_exame_3,
		resultado_exame3 = resultado_exame3,
		data_inicio_tratamento1 = data_inicio_tratamento1,
		droga_administrada1 = droga_administrada1,
		esquema_terapeutico1 = esquema_terapeutico1,
		data_inicio_tratamento2 = data_inicio_tratamento2,
		droga_administrada2 = droga_administrada2,
		esquema_terapeutico2 = esquema_terapeutico2,
		data_inicio_tratamento3 = data_inicio_tratamento3,
		droga_administrada3 = droga_administrada3,
		esquema_terapeutico3 = esquema_terapeutico3,
		hospitalizacao = hospitalizacao,
		data_internacao = data_internacao,
		data_da_alta = data_da_alta,
		uf_hospitalizacao = uf_hospitalizacao,
		municipio_hospitalizacao = municipio_hospitalizacao,
		codigo_ibge_hospitalizacao = codigo_ibge_hospitalizacao,
		nome_hospital_hospitalizacao = nome_hospital_hospitalizacao,
		classificacao_final = classificacao_final,
		criterio_confirmacao = criterio_confirmacao,
		caso_autoctone_municipio_residencia = caso_autoctone_municipio_residencia,
		uf_caso_autoctone = uf_caso_autoctone,
		pais_caso_autoctone = pais_caso_autoctone,
		municipio_caso_autoctone = municipio_caso_autoctone,
		codigo_ibge_caso_autoctone = codigo_ibge_caso_autoctone,
		distrito_caso_autoctone = distrito_caso_autoctone,
		bairro_caso_autoctone = bairro_caso_autoctone,
		area_provavel_infeccao_caso_autoctone = area_provavel_infeccao_caso_autoctone,
		ambiente_infeccao_caso_autoctone = ambiente_infeccao_caso_autoctone,
		doenca_rel_trabalho_caso_autoctone = doenca_rel_trabalho_caso_autoctone,
		evolucao_caso = evolucao_caso,
		status_caso = status_caso,
		data_obito = data_obito,
		data_encerramento = data_encerramento,
		observacao = observacao,
		nome_investigador = nome_investigador,
		funcao_investigador = funcao_investigador,
		email_investigador = email_investigador,
		telefone_investigador = telefone_investigador,
		conselho_classe_investigador = conselho_classe_investigador

		)


	return redirect('my_datas')

@login_required(login_url='/login/')
def export_casos_cancelados(request):
	if request.user.funcao == 'admin':
		casos_cancelados = CasoEsporotricose.objects.filter(status_caso='Cancelado')
		filtro_data_inicio = request.POST.get('filtro_data_inicio')
		filtro_data_fim = request.POST.get('filtro_data_fim')
		if filtro_data_inicio:
			if filtro_data_fim:
				casos = casos_cancelados.filter(data_notificacao__range=[filtro_data_inicio, filtro_data_fim])
			else:
				casos = casos_cancelados.filter(data_notificacao__range=[filtro_data_inicio, datetime.now()])
		else:
			casos = casos_cancelados
		df = pd.DataFrame(list(casos.order_by('-data_notificacao').values()))
		try:
			for i in df.municipio:
				try: # Checando se o valor é diferente de NaN
					i = int(i)
				except:
					continue
				else: # Buscando no modelo municipio o nome de municipio pelo id e alterando o dataframe
					municipio = Municipio.objects.get(id=i)
					df.municipio = df.municipio.replace([i], municipio.nome)
		except AttributeError:
			pass
		response = HttpResponse(content_type='text/csv')
		response['Content-Disposition'] = 'attachment; filename=casos_cancelados.csv'
		df.to_csv(response, index=False)
		
		return response
	else:
		return redirect('casos_cancelados')

	

@login_required(login_url='/login/')
def export_data_csv(request):

	# Pegando todos os casos registrados
	casos = CasoEsporotricose.objects.all().values()
	
	# Filtrando se houver filtro.
	filtro_data_inicio = request.GET.get('filtro_data_inicio')
	filtro_data_fim = request.GET.get('filtro_data_fim')
	filtros_data = [filtro_data_inicio, filtro_data_fim]
	
	for filtro in filtros_data:
		if filtro == '' or filtro == None:
			filtros_data.remove(filtro)
	
	if len(filtros_data) == 2:
		casos_filtrados = casos.filter(data_primeiros_sintomas__range=[filtro_data_inicio,filtro_data_fim]).order_by('-id')
	
	elif len(filtros_data) == 1 and filtros_data[0] != '':
		filtro_unico_dia = filtros_data[0]
		casos_filtrados = casos.filter(data_primeiros_sintomas=filtro_unico_dia).order_by('-id')
	
	else:
		casos_filtrados = casos.order_by('-id')
	
	# Filtrando o tipo de perfil para limitar os casos.
	if request.user.funcao == 'autocadastro':
		# Perfil Auto-Cadastro
		auto_cadastro_id = request.user.id
		casos_response = casos_filtrados.filter(responsavel_pelas_informacoes_id=auto_cadastro_id).order_by('-id')

	elif request.user.funcao == 'municipal':
		# Perfil Municipal
		user_municipio_id = request.user.municipio_id
		user_municipio_nome = str(Municipio.objects.filter(id=user_municipio_id)[0]).upper()
		casos_response = casos_filtrados.filter(Q(municipio_residencia=user_municipio_id) | 
			Q(municipio_residencia=user_municipio_nome) | Q(responsavel_pelas_informacoes_id=user_municipio_id)).order_by('-id')

	elif request.user.funcao == 'gerencia_regional':
		# Perfil Gerencia Regional
		user_gerencia_regional = Municipio.objects.get(id=request.user.municipio_id).gerencia_id
		casos_response = casos_filtrados.filter(gerencia_id=user_gerencia_regional).order_by('-id')
	
	else: # Qualquer outro tipo de perfil
		casos_response = casos_filtrados
		
	# Convertendo em dataframe e alterando o campo município.
	try:
		df = pd.DataFrame(list(casos_response.order_by('-id')))
		for i in df.municipio:
			
			try: # Checando se o valor é diferente de NaN
				i = int(i)
			except:
				continue
			else: # Buscando no modelo municipio o nome de municipio pelo id e alterando o dataframe
				municipio = Municipio.objects.get(id=i)
				df.municipio = df.municipio.replace([i], municipio.nome)
	
	except:
		return redirect('/my_datas/')

	else:
		# Escrevendo o excel e enviando o response.
		with BytesIO() as b:
			
			writer = pd.ExcelWriter(b, engine='openpyxl')
			df.to_excel(writer, sheet_name='Sheet1', index=False)
			writer.save()
			
			filename = 'casos_esporotricose_humana.xlsx'
			response = HttpResponse(b.getvalue(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			response['Content-Disposition'] = 'attachment; filename=%s' % filename
			
			return response
		
	
	'''
	#user = request.user
	municipio_user_logado = request.user.municipio.id
	#print("municipio do usuário:", municipio_user_logado)
	municipio_user_ = Municipio.objects.get(id=municipio_user_logado)
	municipio_user = municipio_user_.id
	#print("id do municipio do user:", municipio_user_)
	#municipio_user = municipio_user_.id
	#print("id do municipio do user:", municipio_user)
	#municipio_user = Municipio.objects.get(id=municipio_user_id)
	casos_ = CasoEsporotricose.objects.all()

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

	response['Content-Disposition'] = 'attachment; filename=myfile.xlsx'

	workbook = Workbook()
	worksheet = workbook.active

	columns = [		
		'tipo_notificacao','agravo_doenca','codigo_cib10','data_notificacao','estado','municipio','codigo_ibge',
		'data_primeiros_sintomas','unidade_saude','unidade_saude_outro','nome_paciente','data_nascimento_paciente',
		'idade_paciente','sexo_paciente','paciente_gestante','raca_paciente','escolaridade_paciente','cartao_sus_paciente',
		'nome_mae_paciente','cep_residencia','uf_residencia','municipio_residencia','bairro_residencia',
		'codigo_ibge_residencia','rua_residencia','numero_residencia','complemento_residencia','distrito_residencia',
		'ponto_referencia_residencia','telefone_residencia','zona_residencia','pais_residencia','data_investigacao',
		'ocupacao','ambientes_frequentados','animais_que_teve_contato','natureza_contato_animais','relacao_animal_doente',
		'exerce_atividade_contato_plantas','historico_contato_material','presenca_lesao_pele','natureza_lesao',
		'natureza_lesao_outro','local_lesao','local_lesao_outro','diagnostico_forma_extrac_doenca',
		'localizacao_forma_extrac_doenca','houve_coleta_material','data_coleta1','numero_gal1','data_coleta2',
		'numero_gal2','data_coleta3','numero_gal3','resultado_isolamento','agente','histopatologia','data_resultado_exame1',
		'descricao_exame_1','resultado_exame1','data_resultado_exame2','descricao_exame_2','resultado_exame2',
		'data_resultado_exame3','descricao_exame_3','resultado_exame3','data_inicio_tratamento1','droga_administrada1',
		'esquema_terapeutico1','data_inicio_tratamento2','droga_administrada2','esquema_terapeutico2',
		'data_inicio_tratamento3','droga_administrada3','esquema_terapeutico3','hospitalizacao','data_internacao',
		'data_da_alta','uf_hospitalizacao','municipio_hospitalizacao','codigo_ibge_hospitalizacao',
		'nome_hospital_hospitalizacao','classificacao_final','criterio_confirmacao','caso_autoctone_municipio_residencia',
		'uf_caso_autoctone','pais_caso_autoctone','municipio_caso_autoctone','codigo_ibge_caso_autoctone',
		'distrito_caso_autoctone','bairro_caso_autoctone','area_provavel_infeccao_caso_autoctone',
		'ambiente_infeccao_caso_autoctone','doenca_rel_trabalho_caso_autoctone','evolucao_caso',
		'data_obito','data_encerramento','observacao','nome_investigador','funcao_investigador','email_investigador',
		'telefone_investigador','conselho_classe_investigador','numero_unico'
	]

	row_num = 1

	for col_num, column_title in enumerate(columns, 1):
		cell = worksheet.cell(row=row_num, column=col_num)
		cell.value = column_title

	
	if request.user.funcao == 'autocadastro':
		casos = casos_.filter(municipio=municipio_user)

		for s in casos:
			row_num += 1


			if s.tipo_notificacao != None:
				tipo_notificacao = s.tipo_notificacao
			else:
				tipo_notificacao = ""

			if s.agravo_doenca != None:
				agravo_doenca = s.agravo_doenca
			else:
				agravo_doenca = ""

			if s.codigo_cib10 != None:
				codigo_cib10 = s.codigo_cib10
			else:
				codigo_cib10 = ""

			
			if s.data_notificacao != None:
				data_notificacao = s.data_notificacao
			else:
				data_notificacao = ""

			
			if s.estado != None:
				estado = s.estado
			else:
				estado = ""

			if s.municipio != None:
				municipio = s.municipio
			else:
				municipio = ""

			if s.codigo_ibge != None:
				codigo_ibge = s.codigo_ibge
			else:
				codigo_ibge = ""

			if  s.data_primeiros_sintomas!= None:
				data_primeiros_sintomas = s.data_primeiros_sintomas
			else:
				data_primeiros_sintomas = ""

			if s.unidade_saude != None:
				unidade_saude = s.unidade_saude
			else:
				unidade_saude = ""

			if s.unidade_saude_outro != None:
				unidade_saude_outro = s.unidade_saude_outro
			else:
				unidade_saude_outro = ""

			if s.nome_paciente != None:
				nome_paciente = s.nome_paciente
			else:
				nome_paciente = ""

			if s.data_nascimento_paciente != None:
				data_nascimento_paciente = s.data_nascimento_paciente
			else:
				data_nascimento_paciente = ""

			if s.idade_paciente != None:
				idade_paciente = s.idade_paciente
			else:
				idade_paciente = ""

			if s.sexo_paciente != None:
				sexo_paciente = s.sexo_paciente
			else:
				sexo_paciente = ""

			if s.paciente_gestante != None:
				paciente_gestante = s.paciente_gestante
			else:
				paciente_gestante = ""

			if s.raca_paciente != None:
				raca_paciente = s.raca_paciente
			else:
				raca_paciente = ""

			if s.escolaridade_paciente != None:
				escolaridade_paciente = s.escolaridade_paciente
			else:
				escolaridade_paciente = ""

			if s.cartao_sus_paciente != None:
				cartao_sus_paciente = s.cartao_sus_paciente
			else:
				cartao_sus_paciente = ""

			if s.nome_mae_paciente != None:
				nome_mae_paciente = s.nome_mae_paciente
			else:
				nome_mae_paciente = ""

			if s.cep_residencia != None:
				cep_residencia = s.cep_residencia
			else:
				cep_residencia = ""

			if s.uf_residencia != None:
				uf_residencia = s.uf_residencia
			else:
				uf_residencia = ""

			if s.municipio_residencia != None:
				municipio_residencia = s.municipio_residencia
			else:
				municipio_residencia = ""

			if s.bairro_residencia != None:
				bairro_residencia = s.bairro_residencia
			else:
				bairro_residencia = ""

			if s.codigo_ibge_residencia != None:
				codigo_ibge_residencia = s.codigo_ibge_residencia
			else:
				codigo_ibge_residencia = ""


			if s.rua_residencia != None:
				rua_residencia = s.rua_residencia
			else:
				rua_residencia = ""

			if s.numero_residencia != None:
				numero_residencia = s.numero_residencia
			else:
				numero_residencia = ""

			if s.complemento_residencia != None:
				complemento_residencia = s.complemento_residencia
			else:
				complemento_residencia = ""

			if s.distrito_residencia != None:
				distrito_residencia = s.distrito_residencia
			else:
				distrito_residencia = ""

			if s.ponto_referencia_residencia != None:
				ponto_referencia_residencia = s.ponto_referencia_residencia
			else:
				ponto_referencia_residencia = ""

			if s.telefone_residencia != None:
				telefone_residencia = s.telefone_residencia
			else:
				telefone_residencia = ""

			if s.zona_residencia != None:
				zona_residencia = s.zona_residencia
			else:
				zona_residencia = ""

			if s.pais_residencia != None:
				pais_residencia = s.pais_residencia
			else:
				pais_residencia = ""

			if s.data_investigacao != None:
				data_investigacao = s.data_investigacao
			else:
				data_investigacao = ""

			if s.ocupacao != None:
				ocupacao = s.ocupacao
			else:
				ocupacao = ""

			if s.ambientes_frequentados != None:
				ambientes_frequentados = s.ambientes_frequentados
			else:
				ambientes_frequentados = ""

			if s.animais_que_teve_contato != None:
				animais_que_teve_contato = s.animais_que_teve_contato
			else:
				animais_que_teve_contato = ""

			if s.natureza_contato_animais != None:
				natureza_contato_animais = s.natureza_contato_animais
			else:
				natureza_contato_animais = ""

			if s.relacao_animal_doente != None:
				relacao_animal_doente = s.relacao_animal_doente
			else:
				relacao_animal_doente = ""

			if s.exerce_atividade_contato_plantas != None:
				exerce_atividade_contato_plantas = s.exerce_atividade_contato_plantas
			else:
				exerce_atividade_contato_plantas = ""

			if s.historico_contato_material != None:
				historico_contato_material = s.historico_contato_material
			else:
				historico_contato_material = ""

			if s.presenca_lesao_pele != None:
				presenca_lesao_pele = s.presenca_lesao_pele
			else:
				presenca_lesao_pele = ""

			if s.natureza_lesao != None:
				natureza_lesao = s.natureza_lesao
			else:
				natureza_lesao = ""

			if s.natureza_lesao_outro != None:
				natureza_lesao_outro = s.natureza_lesao_outro
			else:
				natureza_lesao_outro = ""

			if s.local_lesao != None:
				local_lesao = s.local_lesao
			else:
				local_lesao = ""

			if s.local_lesao_outro != None:
				local_lesao_outro = s.local_lesao_outro
			else:
				local_lesao_outro = ""

			if s.diagnostico_forma_extrac_doenca != None:
				diagnostico_forma_extrac_doenca = s.diagnostico_forma_extrac_doenca
			else:
				diagnostico_forma_extrac_doenca = ""

			if s.localizacao_forma_extrac_doenca != None:
				localizacao_forma_extrac_doenca = s.localizacao_forma_extrac_doenca
			else:
				localizacao_forma_extrac_doenca = ""

			if s.houve_coleta_material != None:
				houve_coleta_material = s.houve_coleta_material
			else:
				houve_coleta_material = ""

			if s.data_coleta1 != None:
				data_coleta1 = s.data_coleta1
			else:
				data_coleta1 = ""

			if s.numero_gal1 != None:
				numero_gal1 = s.numero_gal1
			else:
				numero_gal1 = ""

			if s.data_coleta2 != None:
				data_coleta2 = s.data_coleta2
			else:
				data_coleta2 = ""

			if s.numero_gal2 != None:
				numero_gal2 = s.numero_gal2
			else:
				numero_gal2 = ""

			if s.data_coleta3 != None:
				data_coleta3 = s.data_coleta3
			else:
				data_coleta3 = ""

			if s.numero_gal3 != None:
				numero_gal3 = s.numero_gal3
			else:
				numero_gal3 = ""

			if s.resultado_isolamento != None:
				resultado_isolamento = s.resultado_isolamento
			else:
				resultado_isolamento = ""

			if s.agente != None:
				agente = s.agente
			else:
				agente = ""

			if s.histopatologia != None:
				histopatologia = s.histopatologia
			else:
				histopatologia = ""

			if s.data_resultado_exame1 != None:
				data_resultado_exame1 = s.data_resultado_exame1
			else:
				data_resultado_exame1 = ""

			if s.descricao_exame_1 != None:
				descricao_exame_1 = s.descricao_exame_1
			else:
				descricao_exame_1 = ""

			if s.resultado_exame1 != None:
				resultado_exame1 = s.resultado_exame1
			else:
				resultado_exame1 = ""

			if s.data_resultado_exame2 != None:
				data_resultado_exame2 = s.data_resultado_exame2
			else:
				data_resultado_exame2 = ""

			if s.descricao_exame_2 != None:
				descricao_exame_2 = s.descricao_exame_2
			else:
				descricao_exame_2 = ""

			if s.resultado_exame2 != None:
				resultado_exame2 = s.resultado_exame2
			else:
				resultado_exame2 = ""

			if s.data_resultado_exame3 != None:
				data_resultado_exame3 = s.data_resultado_exame3
			else:
				data_resultado_exame3 = ""

			if s.descricao_exame_3 != None:
				descricao_exame_3 = s.descricao_exame_3
			else:
				descricao_exame_3 = ""

			if s.resultado_exame3 != None:
				resultado_exame3 = s.resultado_exame3
			else:
				resultado_exame3 = ""

			if s.data_inicio_tratamento1 != None:
				data_inicio_tratamento1 = s.data_inicio_tratamento1
			else:
				data_inicio_tratamento1 = ""

			if s.droga_administrada1 != None:
				droga_administrada1 = s.droga_administrada1
			else:
				droga_administrada1 = ""

			if s.esquema_terapeutico1 != None:
				esquema_terapeutico1 = s.esquema_terapeutico1
			else:
				esquema_terapeutico1 = ""

			if s.data_inicio_tratamento2 != None:
				data_inicio_tratamento2 = s.data_inicio_tratamento2
			else:
				data_inicio_tratamento2 = ""

			if s.droga_administrada2 != None:
				droga_administrada2 = s.droga_administrada2
			else:
				droga_administrada2 = ""

			if s.esquema_terapeutico2 != None:
				esquema_terapeutico2 = s.esquema_terapeutico2
			else:
				esquema_terapeutico2 = ""

			if s.data_inicio_tratamento3 != None:
				data_inicio_tratamento3 = s.data_inicio_tratamento3
			else:
				data_inicio_tratamento3 = ""

			if s.droga_administrada3 != None:
				droga_administrada3 = s.droga_administrada3
			else:
				droga_administrada3 = ""

			if s.esquema_terapeutico3 != None:
				esquema_terapeutico3 = s.esquema_terapeutico3
			else:
				esquema_terapeutico3 = ""

			if s.hospitalizacao != None:
				hospitalizacao = s.hospitalizacao
			else:
				hospitalizacao = ""

			if s.data_internacao != None:
				data_internacao = s.data_internacao
			else:
				data_internacao = ""

			if s.data_da_alta != None:
				data_da_alta = s.data_da_alta
			else:
				data_da_alta = ""

			if s.uf_hospitalizacao != None:
				uf_hospitalizacao = s.uf_hospitalizacao
			else:
				uf_hospitalizacao = ""

			if s.municipio_hospitalizacao != None:
				municipio_hospitalizacao = s.municipio_hospitalizacao
			else:
				municipio_hospitalizacao = ""

			if s.codigo_ibge_hospitalizacao != None:
				codigo_ibge_hospitalizacao = s.codigo_ibge_hospitalizacao
			else:
				codigo_ibge_hospitalizacao = ""

			if s.nome_hospital_hospitalizacao != None:
				nome_hospital_hospitalizacao = s.nome_hospital_hospitalizacao
			else:
				nome_hospital_hospitalizacao = ""

			if s.classificacao_final != None:
				classificacao_final = s.classificacao_final
			else:
				classificacao_final = ""

			if s.criterio_confirmacao != None:
				criterio_confirmacao = s.criterio_confirmacao
			else:
				criterio_confirmacao = ""

			if s.caso_autoctone_municipio_residencia != None:
				caso_autoctone_municipio_residencia = s.caso_autoctone_municipio_residencia
			else:
				caso_autoctone_municipio_residencia = ""

			if s.uf_caso_autoctone != None:
				uf_caso_autoctone = s.uf_caso_autoctone
			else:
				uf_caso_autoctone = ""

			if s.pais_caso_autoctone != None:
				pais_caso_autoctone = s.pais_caso_autoctone
			else:
				pais_caso_autoctone = ""

			if s.municipio_caso_autoctone != None:
				municipio_caso_autoctone = s.municipio_caso_autoctone
			else:
				municipio_caso_autoctone = ""

			if s.codigo_ibge_caso_autoctone != None:
				codigo_ibge_caso_autoctone = s.codigo_ibge_caso_autoctone
			else:
				codigo_ibge_caso_autoctone = ""

			if s.distrito_caso_autoctone != None:
				distrito_caso_autoctone = s.distrito_caso_autoctone
			else:
				distrito_caso_autoctone = ""

			if s.bairro_caso_autoctone != None:
				bairro_caso_autoctone = s.bairro_caso_autoctone
			else:
				bairro_caso_autoctone = ""

			if s.area_provavel_infeccao_caso_autoctone != None:
				area_provavel_infeccao_caso_autoctone = s.area_provavel_infeccao_caso_autoctone
			else:
				area_provavel_infeccao_caso_autoctone = ""

			if s.ambiente_infeccao_caso_autoctone != None:
				ambiente_infeccao_caso_autoctone = s.ambiente_infeccao_caso_autoctone
			else:
				ambiente_infeccao_caso_autoctone = ""

			if s.doenca_rel_trabalho_caso_autoctone != None:
				doenca_rel_trabalho_caso_autoctone = s.doenca_rel_trabalho_caso_autoctone
			else:
				doenca_rel_trabalho_caso_autoctone = ""

			if s.evolucao_caso != None:
				evolucao_caso = s.evolucao_caso
			else:
				evolucao_caso = ""

			if s.data_obito != None:
				data_obito = s.data_obito
			else:
				data_obito = ""

			if s.data_encerramento != None:
				data_encerramento = s.data_encerramento
			else:
				data_encerramento = ""

			if s.observacao != None:
				observacao = s.observacao
			else:
				observacao = ""

			if s.nome_investigador != None:
				nome_investigador = s.nome_investigador
			else:
				nome_investigador = ""

			if s.funcao_investigador != None:
				funcao_investigador = s.funcao_investigador
			else:
				funcao_investigador = ""

			if s.email_investigador != None:
				email_investigador = s.email_investigador
			else:
				email_investigador = ""

			if s.telefone_investigador != None:
				telefone_investigador = s.telefone_investigador
			else:
				telefone_investigador = ""

			if s.conselho_classe_investigador != None:
				conselho_classe_investigador = s.conselho_classe_investigador
			else:
				conselho_classe_investigador = ""

			if s.numero_unico != None:
				numero_unico = s.numero_unico
			else:
				numero_unico = ""

		

		row = [
			tipo_notificacao,
			agravo_doenca,
			codigo_cib10,
			data_notificacao,
			estado,
			municipio,
			codigo_ibge,
			data_primeiros_sintomas,
			unidade_saude,
			unidade_saude_outro,
			nome_paciente,
			data_nascimento_paciente,
			idade_paciente,
			sexo_paciente,
			paciente_gestante,
			raca_paciente,
			escolaridade_paciente,
			cartao_sus_paciente,
			nome_mae_paciente,
			cep_residencia,
			uf_residencia,
			municipio_residencia,
			bairro_residencia,
			codigo_ibge_residencia,
			rua_residencia,
			numero_residencia,
			complemento_residencia,
			distrito_residencia,
			ponto_referencia_residencia,
			telefone_residencia,
			zona_residencia,
			pais_residencia,
			data_investigacao,
			ocupacao,
			ambientes_frequentados,
			animais_que_teve_contato,
			natureza_contato_animais,
			relacao_animal_doente,
			exerce_atividade_contato_plantas,
			historico_contato_material,
			presenca_lesao_pele,
			natureza_lesao,
			natureza_lesao_outro,
			local_lesao,
			local_lesao_outro,
			diagnostico_forma_extrac_doenca,
			localizacao_forma_extrac_doenca,
			houve_coleta_material,
			data_coleta1,
			numero_gal1,
			data_coleta2,
			numero_gal2,
			data_coleta3,
			numero_gal3,
			resultado_isolamento,
			agente,
			histopatologia,
			data_resultado_exame1,
			descricao_exame_1,
			resultado_exame1,
			data_resultado_exame2,
			descricao_exame_2,
			resultado_exame2,
			data_resultado_exame3,
			descricao_exame_3,
			resultado_exame3,
			data_inicio_tratamento1,
			droga_administrada1,
			esquema_terapeutico1,
			data_inicio_tratamento2,
			droga_administrada2,
			esquema_terapeutico2,
			data_inicio_tratamento3,
			droga_administrada3,
			esquema_terapeutico3,
			hospitalizacao,
			data_internacao,
			data_da_alta,
			uf_hospitalizacao,
			municipio_hospitalizacao,
			codigo_ibge_hospitalizacao,
			nome_hospital_hospitalizacao,
			classificacao_final,
			criterio_confirmacao,
			caso_autoctone_municipio_residencia,
			uf_caso_autoctone,
			pais_caso_autoctone,
			municipio_caso_autoctone,
			codigo_ibge_caso_autoctone,
			distrito_caso_autoctone,
			bairro_caso_autoctone,
			area_provavel_infeccao_caso_autoctone,
			ambiente_infeccao_caso_autoctone,
			doenca_rel_trabalho_caso_autoctone,
			evolucao_caso,
			data_obito,
			data_encerramento,
			observacao,
			nome_investigador,
			funcao_investigador,
			email_investigador,
			telefone_investigador,
			conselho_classe_investigador,
			numero_unico,

		]
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell_value_ = str(cell_value)
			cell.value = cell_value_

		workbook.save(response)

		return response


	elif request.user.funcao == 'municipal':
		casos = casos_.filter(municipio=municipio_user)

		for s in casos:
			row_num += 1


			if s.tipo_notificacao != None:
				tipo_notificacao = s.tipo_notificacao
			else:
				tipo_notificacao = ""

			if s.agravo_doenca != None:
				agravo_doenca = s.agravo_doenca
			else:
				agravo_doenca = ""

			if s.codigo_cib10 != None:
				codigo_cib10 = s.codigo_cib10
			else:
				codigo_cib10 = ""

			
			if s.data_notificacao != None:
				data_notificacao = s.data_notificacao
			else:
				data_notificacao = ""

			
			if s.estado != None:
				estado = s.estado
			else:
				estado = ""

			if s.municipio != None:
				municipio = s.municipio
			else:
				municipio = ""

			if s.codigo_ibge != None:
				codigo_ibge = s.codigo_ibge
			else:
				codigo_ibge = ""

			if  s.data_primeiros_sintomas!= None:
				data_primeiros_sintomas = s.data_primeiros_sintomas
			else:
				data_primeiros_sintomas = ""

			if s.unidade_saude != None:
				unidade_saude = s.unidade_saude
			else:
				unidade_saude = ""

			if s.unidade_saude_outro != None:
				unidade_saude_outro = s.unidade_saude_outro
			else:
				unidade_saude_outro = ""

			if s.nome_paciente != None:
				nome_paciente = s.nome_paciente
			else:
				nome_paciente = ""

			if s.data_nascimento_paciente != None:
				data_nascimento_paciente = s.data_nascimento_paciente
			else:
				data_nascimento_paciente = ""

			if s.idade_paciente != None:
				idade_paciente = s.idade_paciente
			else:
				idade_paciente = ""

			if s.sexo_paciente != None:
				sexo_paciente = s.sexo_paciente
			else:
				sexo_paciente = ""

			if s.paciente_gestante != None:
				paciente_gestante = s.paciente_gestante
			else:
				paciente_gestante = ""

			if s.raca_paciente != None:
				raca_paciente = s.raca_paciente
			else:
				raca_paciente = ""

			if s.escolaridade_paciente != None:
				escolaridade_paciente = s.escolaridade_paciente
			else:
				escolaridade_paciente = ""

			if s.cartao_sus_paciente != None:
				cartao_sus_paciente = s.cartao_sus_paciente
			else:
				cartao_sus_paciente = ""

			if s.nome_mae_paciente != None:
				nome_mae_paciente = s.nome_mae_paciente
			else:
				nome_mae_paciente = ""

			if s.cep_residencia != None:
				cep_residencia = s.cep_residencia
			else:
				cep_residencia = ""

			if s.uf_residencia != None:
				uf_residencia = s.uf_residencia
			else:
				uf_residencia = ""

			if s.municipio_residencia != None:
				municipio_residencia = s.municipio_residencia
			else:
				municipio_residencia = ""

			if s.bairro_residencia != None:
				bairro_residencia = s.bairro_residencia
			else:
				bairro_residencia = ""

			if s.codigo_ibge_residencia != None:
				codigo_ibge_residencia = s.codigo_ibge_residencia
			else:
				codigo_ibge_residencia = ""


			if s.rua_residencia != None:
				rua_residencia = s.rua_residencia
			else:
				rua_residencia = ""

			if s.numero_residencia != None:
				numero_residencia = s.numero_residencia
			else:
				numero_residencia = ""

			if s.complemento_residencia != None:
				complemento_residencia = s.complemento_residencia
			else:
				complemento_residencia = ""

			if s.distrito_residencia != None:
				distrito_residencia = s.distrito_residencia
			else:
				distrito_residencia = ""

			if s.ponto_referencia_residencia != None:
				ponto_referencia_residencia = s.ponto_referencia_residencia
			else:
				ponto_referencia_residencia = ""

			if s.telefone_residencia != None:
				telefone_residencia = s.telefone_residencia
			else:
				telefone_residencia = ""

			if s.zona_residencia != None:
				zona_residencia = s.zona_residencia
			else:
				zona_residencia = ""

			if s.pais_residencia != None:
				pais_residencia = s.pais_residencia
			else:
				pais_residencia = ""

			if s.data_investigacao != None:
				data_investigacao = s.data_investigacao
			else:
				data_investigacao = ""

			if s.ocupacao != None:
				ocupacao = s.ocupacao
			else:
				ocupacao = ""

			if s.ambientes_frequentados != None:
				ambientes_frequentados = s.ambientes_frequentados
			else:
				ambientes_frequentados = ""

			if s.animais_que_teve_contato != None:
				animais_que_teve_contato = s.animais_que_teve_contato
			else:
				animais_que_teve_contato = ""

			if s.natureza_contato_animais != None:
				natureza_contato_animais = s.natureza_contato_animais
			else:
				natureza_contato_animais = ""

			if s.relacao_animal_doente != None:
				relacao_animal_doente = s.relacao_animal_doente
			else:
				relacao_animal_doente = ""

			if s.exerce_atividade_contato_plantas != None:
				exerce_atividade_contato_plantas = s.exerce_atividade_contato_plantas
			else:
				exerce_atividade_contato_plantas = ""

			if s.historico_contato_material != None:
				historico_contato_material = s.historico_contato_material
			else:
				historico_contato_material = ""

			if s.presenca_lesao_pele != None:
				presenca_lesao_pele = s.presenca_lesao_pele
			else:
				presenca_lesao_pele = ""

			if s.natureza_lesao != None:
				natureza_lesao = s.natureza_lesao
			else:
				natureza_lesao = ""

			if s.natureza_lesao_outro != None:
				natureza_lesao_outro = s.natureza_lesao_outro
			else:
				natureza_lesao_outro = ""

			if s.local_lesao != None:
				local_lesao = s.local_lesao
			else:
				local_lesao = ""

			if s.local_lesao_outro != None:
				local_lesao_outro = s.local_lesao_outro
			else:
				local_lesao_outro = ""

			if s.diagnostico_forma_extrac_doenca != None:
				diagnostico_forma_extrac_doenca = s.diagnostico_forma_extrac_doenca
			else:
				diagnostico_forma_extrac_doenca = ""

			if s.localizacao_forma_extrac_doenca != None:
				localizacao_forma_extrac_doenca = s.localizacao_forma_extrac_doenca
			else:
				localizacao_forma_extrac_doenca = ""

			if s.houve_coleta_material != None:
				houve_coleta_material = s.houve_coleta_material
			else:
				houve_coleta_material = ""

			if s.data_coleta1 != None:
				data_coleta1 = s.data_coleta1
			else:
				data_coleta1 = ""

			if s.numero_gal1 != None:
				numero_gal1 = s.numero_gal1
			else:
				numero_gal1 = ""

			if s.data_coleta2 != None:
				data_coleta2 = s.data_coleta2
			else:
				data_coleta2 = ""

			if s.numero_gal2 != None:
				numero_gal2 = s.numero_gal2
			else:
				numero_gal2 = ""

			if s.data_coleta3 != None:
				data_coleta3 = s.data_coleta3
			else:
				data_coleta3 = ""

			if s.numero_gal3 != None:
				numero_gal3 = s.numero_gal3
			else:
				numero_gal3 = ""

			if s.resultado_isolamento != None:
				resultado_isolamento = s.resultado_isolamento
			else:
				resultado_isolamento = ""

			if s.agente != None:
				agente = s.agente
			else:
				agente = ""

			if s.histopatologia != None:
				histopatologia = s.histopatologia
			else:
				histopatologia = ""

			if s.data_resultado_exame1 != None:
				data_resultado_exame1 = s.data_resultado_exame1
			else:
				data_resultado_exame1 = ""

			if s.descricao_exame_1 != None:
				descricao_exame_1 = s.descricao_exame_1
			else:
				descricao_exame_1 = ""

			if s.resultado_exame1 != None:
				resultado_exame1 = s.resultado_exame1
			else:
				resultado_exame1 = ""

			if s.data_resultado_exame2 != None:
				data_resultado_exame2 = s.data_resultado_exame2
			else:
				data_resultado_exame2 = ""

			if s.descricao_exame_2 != None:
				descricao_exame_2 = s.descricao_exame_2
			else:
				descricao_exame_2 = ""

			if s.resultado_exame2 != None:
				resultado_exame2 = s.resultado_exame2
			else:
				resultado_exame2 = ""

			if s.data_resultado_exame3 != None:
				data_resultado_exame3 = s.data_resultado_exame3
			else:
				data_resultado_exame3 = ""

			if s.descricao_exame_3 != None:
				descricao_exame_3 = s.descricao_exame_3
			else:
				descricao_exame_3 = ""

			if s.resultado_exame3 != None:
				resultado_exame3 = s.resultado_exame3
			else:
				resultado_exame3 = ""

			if s.data_inicio_tratamento1 != None:
				data_inicio_tratamento1 = s.data_inicio_tratamento1
			else:
				data_inicio_tratamento1 = ""

			if s.droga_administrada1 != None:
				droga_administrada1 = s.droga_administrada1
			else:
				droga_administrada1 = ""

			if s.esquema_terapeutico1 != None:
				esquema_terapeutico1 = s.esquema_terapeutico1
			else:
				esquema_terapeutico1 = ""

			if s.data_inicio_tratamento2 != None:
				data_inicio_tratamento2 = s.data_inicio_tratamento2
			else:
				data_inicio_tratamento2 = ""

			if s.droga_administrada2 != None:
				droga_administrada2 = s.droga_administrada2
			else:
				droga_administrada2 = ""

			if s.esquema_terapeutico2 != None:
				esquema_terapeutico2 = s.esquema_terapeutico2
			else:
				esquema_terapeutico2 = ""

			if s.data_inicio_tratamento3 != None:
				data_inicio_tratamento3 = s.data_inicio_tratamento3
			else:
				data_inicio_tratamento3 = ""

			if s.droga_administrada3 != None:
				droga_administrada3 = s.droga_administrada3
			else:
				droga_administrada3 = ""

			if s.esquema_terapeutico3 != None:
				esquema_terapeutico3 = s.esquema_terapeutico3
			else:
				esquema_terapeutico3 = ""

			if s.hospitalizacao != None:
				hospitalizacao = s.hospitalizacao
			else:
				hospitalizacao = ""

			if s.data_internacao != None:
				data_internacao = s.data_internacao
			else:
				data_internacao = ""

			if s.data_da_alta != None:
				data_da_alta = s.data_da_alta
			else:
				data_da_alta = ""

			if s.uf_hospitalizacao != None:
				uf_hospitalizacao = s.uf_hospitalizacao
			else:
				uf_hospitalizacao = ""

			if s.municipio_hospitalizacao != None:
				municipio_hospitalizacao = s.municipio_hospitalizacao
			else:
				municipio_hospitalizacao = ""

			if s.codigo_ibge_hospitalizacao != None:
				codigo_ibge_hospitalizacao = s.codigo_ibge_hospitalizacao
			else:
				codigo_ibge_hospitalizacao = ""

			if s.nome_hospital_hospitalizacao != None:
				nome_hospital_hospitalizacao = s.nome_hospital_hospitalizacao
			else:
				nome_hospital_hospitalizacao = ""

			if s.classificacao_final != None:
				classificacao_final = s.classificacao_final
			else:
				classificacao_final = ""

			if s.criterio_confirmacao != None:
				criterio_confirmacao = s.criterio_confirmacao
			else:
				criterio_confirmacao = ""

			if s.caso_autoctone_municipio_residencia != None:
				caso_autoctone_municipio_residencia = s.caso_autoctone_municipio_residencia
			else:
				caso_autoctone_municipio_residencia = ""

			if s.uf_caso_autoctone != None:
				uf_caso_autoctone = s.uf_caso_autoctone
			else:
				uf_caso_autoctone = ""

			if s.pais_caso_autoctone != None:
				pais_caso_autoctone = s.pais_caso_autoctone
			else:
				pais_caso_autoctone = ""

			if s.municipio_caso_autoctone != None:
				municipio_caso_autoctone = s.municipio_caso_autoctone
			else:
				municipio_caso_autoctone = ""

			if s.codigo_ibge_caso_autoctone != None:
				codigo_ibge_caso_autoctone = s.codigo_ibge_caso_autoctone
			else:
				codigo_ibge_caso_autoctone = ""

			if s.distrito_caso_autoctone != None:
				distrito_caso_autoctone = s.distrito_caso_autoctone
			else:
				distrito_caso_autoctone = ""

			if s.bairro_caso_autoctone != None:
				bairro_caso_autoctone = s.bairro_caso_autoctone
			else:
				bairro_caso_autoctone = ""

			if s.area_provavel_infeccao_caso_autoctone != None:
				area_provavel_infeccao_caso_autoctone = s.area_provavel_infeccao_caso_autoctone
			else:
				area_provavel_infeccao_caso_autoctone = ""

			if s.ambiente_infeccao_caso_autoctone != None:
				ambiente_infeccao_caso_autoctone = s.ambiente_infeccao_caso_autoctone
			else:
				ambiente_infeccao_caso_autoctone = ""

			if s.doenca_rel_trabalho_caso_autoctone != None:
				doenca_rel_trabalho_caso_autoctone = s.doenca_rel_trabalho_caso_autoctone
			else:
				doenca_rel_trabalho_caso_autoctone = ""

			if s.evolucao_caso != None:
				evolucao_caso = s.evolucao_caso
			else:
				evolucao_caso = ""

			if s.data_obito != None:
				data_obito = s.data_obito
			else:
				data_obito = ""

			if s.data_encerramento != None:
				data_encerramento = s.data_encerramento
			else:
				data_encerramento = ""

			if s.observacao != None:
				observacao = s.observacao
			else:
				observacao = ""

			if s.nome_investigador != None:
				nome_investigador = s.nome_investigador
			else:
				nome_investigador = ""

			if s.funcao_investigador != None:
				funcao_investigador = s.funcao_investigador
			else:
				funcao_investigador = ""

			if s.email_investigador != None:
				email_investigador = s.email_investigador
			else:
				email_investigador = ""

			if s.telefone_investigador != None:
				telefone_investigador = s.telefone_investigador
			else:
				telefone_investigador = ""

			if s.conselho_classe_investigador != None:
				conselho_classe_investigador = s.conselho_classe_investigador
			else:
				conselho_classe_investigador = ""

			if s.numero_unico != None:
				numero_unico = s.numero_unico
			else:
				numero_unico = ""

		

			row = [
				tipo_notificacao,
				agravo_doenca,
				codigo_cib10,
				data_notificacao,
				estado,
				municipio,
				codigo_ibge,
				data_primeiros_sintomas,
				unidade_saude,
				unidade_saude_outro,
				nome_paciente,
				data_nascimento_paciente,
				idade_paciente,
				sexo_paciente,
				paciente_gestante,
				raca_paciente,
				escolaridade_paciente,
				cartao_sus_paciente,
				nome_mae_paciente,
				cep_residencia,
				uf_residencia,
				municipio_residencia,
				bairro_residencia,
				codigo_ibge_residencia,
				rua_residencia,
				numero_residencia,
				complemento_residencia,
				distrito_residencia,
				ponto_referencia_residencia,
				telefone_residencia,
				zona_residencia,
				pais_residencia,
				data_investigacao,
				ocupacao,
				ambientes_frequentados,
				animais_que_teve_contato,
				natureza_contato_animais,
				relacao_animal_doente,
				exerce_atividade_contato_plantas,
				historico_contato_material,
				presenca_lesao_pele,
				natureza_lesao,
				natureza_lesao_outro,
				local_lesao,
				local_lesao_outro,
				diagnostico_forma_extrac_doenca,
				localizacao_forma_extrac_doenca,
				houve_coleta_material,
				data_coleta1,
				numero_gal1,
				data_coleta2,
				numero_gal2,
				data_coleta3,
				numero_gal3,
				resultado_isolamento,
				agente,
				histopatologia,
				data_resultado_exame1,
				descricao_exame_1,
				resultado_exame1,
				data_resultado_exame2,
				descricao_exame_2,
				resultado_exame2,
				data_resultado_exame3,
				descricao_exame_3,
				resultado_exame3,
				data_inicio_tratamento1,
				droga_administrada1,
				esquema_terapeutico1,
				data_inicio_tratamento2,
				droga_administrada2,
				esquema_terapeutico2,
				data_inicio_tratamento3,
				droga_administrada3,
				esquema_terapeutico3,
				hospitalizacao,
				data_internacao,
				data_da_alta,
				uf_hospitalizacao,
				municipio_hospitalizacao,
				codigo_ibge_hospitalizacao,
				nome_hospital_hospitalizacao,
				classificacao_final,
				criterio_confirmacao,
				caso_autoctone_municipio_residencia,
				uf_caso_autoctone,
				pais_caso_autoctone,
				municipio_caso_autoctone,
				codigo_ibge_caso_autoctone,
				distrito_caso_autoctone,
				bairro_caso_autoctone,
				area_provavel_infeccao_caso_autoctone,
				ambiente_infeccao_caso_autoctone,
				doenca_rel_trabalho_caso_autoctone,
				evolucao_caso,
				data_obito,
				data_encerramento,
				observacao,
				nome_investigador,
				funcao_investigador,
				email_investigador,
				telefone_investigador,
				conselho_classe_investigador,
				numero_unico,

			]
			for col_num, cell_value in enumerate(row, 1):
				cell = worksheet.cell(row=row_num, column=col_num)
				cell_value_ = str(cell_value)
				cell.value = cell_value_

			workbook.save(response)

			return response


	else:
		casos = casos_

		for s in casos:
			row_num += 1


			if s.tipo_notificacao != None:
				tipo_notificacao = s.tipo_notificacao
			else:
				tipo_notificacao = ""

			if s.agravo_doenca != None:
				agravo_doenca = s.agravo_doenca
			else:
				agravo_doenca = ""

			if s.codigo_cib10 != None:
				codigo_cib10 = s.codigo_cib10
			else:
				codigo_cib10 = ""

			
			if s.data_notificacao != None:
				data_notificacao = s.data_notificacao
			else:
				data_notificacao = ""

			
			if s.estado != None:
				estado = s.estado
			else:
				estado = ""

			if s.municipio != None:
				municipio = s.municipio
			else:
				municipio = ""

			if s.codigo_ibge != None:
				codigo_ibge = s.codigo_ibge
			else:
				codigo_ibge = ""

			if  s.data_primeiros_sintomas!= None:
				data_primeiros_sintomas = s.data_primeiros_sintomas
			else:
				data_primeiros_sintomas = ""

			if s.unidade_saude != None:
				unidade_saude = s.unidade_saude
			else:
				unidade_saude = ""

			if s.unidade_saude_outro != None:
				unidade_saude_outro = s.unidade_saude_outro
			else:
				unidade_saude_outro = ""

			if s.nome_paciente != None:
				nome_paciente = s.nome_paciente
			else:
				nome_paciente = ""

			if s.data_nascimento_paciente != None:
				data_nascimento_paciente = s.data_nascimento_paciente
			else:
				data_nascimento_paciente = ""

			if s.idade_paciente != None:
				idade_paciente = s.idade_paciente
			else:
				idade_paciente = ""

			if s.sexo_paciente != None:
				sexo_paciente = s.sexo_paciente
			else:
				sexo_paciente = ""

			if s.paciente_gestante != None:
				paciente_gestante = s.paciente_gestante
			else:
				paciente_gestante = ""

			if s.raca_paciente != None:
				raca_paciente = s.raca_paciente
			else:
				raca_paciente = ""

			if s.escolaridade_paciente != None:
				escolaridade_paciente = s.escolaridade_paciente
			else:
				escolaridade_paciente = ""

			if s.cartao_sus_paciente != None:
				cartao_sus_paciente = s.cartao_sus_paciente
			else:
				cartao_sus_paciente = ""

			if s.nome_mae_paciente != None:
				nome_mae_paciente = s.nome_mae_paciente
			else:
				nome_mae_paciente = ""

			if s.cep_residencia != None:
				cep_residencia = s.cep_residencia
			else:
				cep_residencia = ""

			if s.uf_residencia != None:
				uf_residencia = s.uf_residencia
			else:
				uf_residencia = ""

			if s.municipio_residencia != None:
				municipio_residencia = s.municipio_residencia
			else:
				municipio_residencia = ""

			if s.bairro_residencia != None:
				bairro_residencia = s.bairro_residencia
			else:
				bairro_residencia = ""

			if s.codigo_ibge_residencia != None:
				codigo_ibge_residencia = s.codigo_ibge_residencia
			else:
				codigo_ibge_residencia = ""


			if s.rua_residencia != None:
				rua_residencia = s.rua_residencia
			else:
				rua_residencia = ""

			if s.numero_residencia != None:
				numero_residencia = s.numero_residencia
			else:
				numero_residencia = ""

			if s.complemento_residencia != None:
				complemento_residencia = s.complemento_residencia
			else:
				complemento_residencia = ""

			if s.distrito_residencia != None:
				distrito_residencia = s.distrito_residencia
			else:
				distrito_residencia = ""

			if s.ponto_referencia_residencia != None:
				ponto_referencia_residencia = s.ponto_referencia_residencia
			else:
				ponto_referencia_residencia = ""

			if s.telefone_residencia != None:
				telefone_residencia = s.telefone_residencia
			else:
				telefone_residencia = ""

			if s.zona_residencia != None:
				zona_residencia = s.zona_residencia
			else:
				zona_residencia = ""

			if s.pais_residencia != None:
				pais_residencia = s.pais_residencia
			else:
				pais_residencia = ""

			if s.data_investigacao != None:
				data_investigacao = s.data_investigacao
			else:
				data_investigacao = ""

			if s.ocupacao != None:
				ocupacao = s.ocupacao
			else:
				ocupacao = ""

			if s.ambientes_frequentados != None:
				ambientes_frequentados = s.ambientes_frequentados
			else:
				ambientes_frequentados = ""

			if s.animais_que_teve_contato != None:
				animais_que_teve_contato = s.animais_que_teve_contato
			else:
				animais_que_teve_contato = ""

			if s.natureza_contato_animais != None:
				natureza_contato_animais = s.natureza_contato_animais
			else:
				natureza_contato_animais = ""

			if s.relacao_animal_doente != None:
				relacao_animal_doente = s.relacao_animal_doente
			else:
				relacao_animal_doente = ""

			if s.exerce_atividade_contato_plantas != None:
				exerce_atividade_contato_plantas = s.exerce_atividade_contato_plantas
			else:
				exerce_atividade_contato_plantas = ""

			if s.historico_contato_material != None:
				historico_contato_material = s.historico_contato_material
			else:
				historico_contato_material = ""

			if s.presenca_lesao_pele != None:
				presenca_lesao_pele = s.presenca_lesao_pele
			else:
				presenca_lesao_pele = ""

			if s.natureza_lesao != None:
				natureza_lesao = s.natureza_lesao
			else:
				natureza_lesao = ""

			if s.natureza_lesao_outro != None:
				natureza_lesao_outro = s.natureza_lesao_outro
			else:
				natureza_lesao_outro = ""

			if s.local_lesao != None:
				local_lesao = s.local_lesao
			else:
				local_lesao = ""

			if s.local_lesao_outro != None:
				local_lesao_outro = s.local_lesao_outro
			else:
				local_lesao_outro = ""

			if s.diagnostico_forma_extrac_doenca != None:
				diagnostico_forma_extrac_doenca = s.diagnostico_forma_extrac_doenca
			else:
				diagnostico_forma_extrac_doenca = ""

			if s.localizacao_forma_extrac_doenca != None:
				localizacao_forma_extrac_doenca = s.localizacao_forma_extrac_doenca
			else:
				localizacao_forma_extrac_doenca = ""

			if s.houve_coleta_material != None:
				houve_coleta_material = s.houve_coleta_material
			else:
				houve_coleta_material = ""

			if s.data_coleta1 != None:
				data_coleta1 = s.data_coleta1
			else:
				data_coleta1 = ""

			if s.numero_gal1 != None:
				numero_gal1 = s.numero_gal1
			else:
				numero_gal1 = ""

			if s.data_coleta2 != None:
				data_coleta2 = s.data_coleta2
			else:
				data_coleta2 = ""

			if s.numero_gal2 != None:
				numero_gal2 = s.numero_gal2
			else:
				numero_gal2 = ""

			if s.data_coleta3 != None:
				data_coleta3 = s.data_coleta3
			else:
				data_coleta3 = ""

			if s.numero_gal3 != None:
				numero_gal3 = s.numero_gal3
			else:
				numero_gal3 = ""

			if s.resultado_isolamento != None:
				resultado_isolamento = s.resultado_isolamento
			else:
				resultado_isolamento = ""

			if s.agente != None:
				agente = s.agente
			else:
				agente = ""

			if s.histopatologia != None:
				histopatologia = s.histopatologia
			else:
				histopatologia = ""

			if s.data_resultado_exame1 != None:
				data_resultado_exame1 = s.data_resultado_exame1
			else:
				data_resultado_exame1 = ""

			if s.descricao_exame_1 != None:
				descricao_exame_1 = s.descricao_exame_1
			else:
				descricao_exame_1 = ""

			if s.resultado_exame1 != None:
				resultado_exame1 = s.resultado_exame1
			else:
				resultado_exame1 = ""

			if s.data_resultado_exame2 != None:
				data_resultado_exame2 = s.data_resultado_exame2
			else:
				data_resultado_exame2 = ""

			if s.descricao_exame_2 != None:
				descricao_exame_2 = s.descricao_exame_2
			else:
				descricao_exame_2 = ""

			if s.resultado_exame2 != None:
				resultado_exame2 = s.resultado_exame2
			else:
				resultado_exame2 = ""

			if s.data_resultado_exame3 != None:
				data_resultado_exame3 = s.data_resultado_exame3
			else:
				data_resultado_exame3 = ""

			if s.descricao_exame_3 != None:
				descricao_exame_3 = s.descricao_exame_3
			else:
				descricao_exame_3 = ""

			if s.resultado_exame3 != None:
				resultado_exame3 = s.resultado_exame3
			else:
				resultado_exame3 = ""

			if s.data_inicio_tratamento1 != None:
				data_inicio_tratamento1 = s.data_inicio_tratamento1
			else:
				data_inicio_tratamento1 = ""

			if s.droga_administrada1 != None:
				droga_administrada1 = s.droga_administrada1
			else:
				droga_administrada1 = ""

			if s.esquema_terapeutico1 != None:
				esquema_terapeutico1 = s.esquema_terapeutico1
			else:
				esquema_terapeutico1 = ""

			if s.data_inicio_tratamento2 != None:
				data_inicio_tratamento2 = s.data_inicio_tratamento2
			else:
				data_inicio_tratamento2 = ""

			if s.droga_administrada2 != None:
				droga_administrada2 = s.droga_administrada2
			else:
				droga_administrada2 = ""

			if s.esquema_terapeutico2 != None:
				esquema_terapeutico2 = s.esquema_terapeutico2
			else:
				esquema_terapeutico2 = ""

			if s.data_inicio_tratamento3 != None:
				data_inicio_tratamento3 = s.data_inicio_tratamento3
			else:
				data_inicio_tratamento3 = ""

			if s.droga_administrada3 != None:
				droga_administrada3 = s.droga_administrada3
			else:
				droga_administrada3 = ""

			if s.esquema_terapeutico3 != None:
				esquema_terapeutico3 = s.esquema_terapeutico3
			else:
				esquema_terapeutico3 = ""

			if s.hospitalizacao != None:
				hospitalizacao = s.hospitalizacao
			else:
				hospitalizacao = ""

			if s.data_internacao != None:
				data_internacao = s.data_internacao
			else:
				data_internacao = ""

			if s.data_da_alta != None:
				data_da_alta = s.data_da_alta
			else:
				data_da_alta = ""

			if s.uf_hospitalizacao != None:
				uf_hospitalizacao = s.uf_hospitalizacao
			else:
				uf_hospitalizacao = ""

			if s.municipio_hospitalizacao != None:
				municipio_hospitalizacao = s.municipio_hospitalizacao
			else:
				municipio_hospitalizacao = ""

			if s.codigo_ibge_hospitalizacao != None:
				codigo_ibge_hospitalizacao = s.codigo_ibge_hospitalizacao
			else:
				codigo_ibge_hospitalizacao = ""

			if s.nome_hospital_hospitalizacao != None:
				nome_hospital_hospitalizacao = s.nome_hospital_hospitalizacao
			else:
				nome_hospital_hospitalizacao = ""

			if s.classificacao_final != None:
				classificacao_final = s.classificacao_final
			else:
				classificacao_final = ""

			if s.criterio_confirmacao != None:
				criterio_confirmacao = s.criterio_confirmacao
			else:
				criterio_confirmacao = ""

			if s.caso_autoctone_municipio_residencia != None:
				caso_autoctone_municipio_residencia = s.caso_autoctone_municipio_residencia
			else:
				caso_autoctone_municipio_residencia = ""

			if s.uf_caso_autoctone != None:
				uf_caso_autoctone = s.uf_caso_autoctone
			else:
				uf_caso_autoctone = ""

			if s.pais_caso_autoctone != None:
				pais_caso_autoctone = s.pais_caso_autoctone
			else:
				pais_caso_autoctone = ""

			if s.municipio_caso_autoctone != None:
				municipio_caso_autoctone = s.municipio_caso_autoctone
			else:
				municipio_caso_autoctone = ""

			if s.codigo_ibge_caso_autoctone != None:
				codigo_ibge_caso_autoctone = s.codigo_ibge_caso_autoctone
			else:
				codigo_ibge_caso_autoctone = ""

			if s.distrito_caso_autoctone != None:
				distrito_caso_autoctone = s.distrito_caso_autoctone
			else:
				distrito_caso_autoctone = ""

			if s.bairro_caso_autoctone != None:
				bairro_caso_autoctone = s.bairro_caso_autoctone
			else:
				bairro_caso_autoctone = ""

			if s.area_provavel_infeccao_caso_autoctone != None:
				area_provavel_infeccao_caso_autoctone = s.area_provavel_infeccao_caso_autoctone
			else:
				area_provavel_infeccao_caso_autoctone = ""

			if s.ambiente_infeccao_caso_autoctone != None:
				ambiente_infeccao_caso_autoctone = s.ambiente_infeccao_caso_autoctone
			else:
				ambiente_infeccao_caso_autoctone = ""

			if s.doenca_rel_trabalho_caso_autoctone != None:
				doenca_rel_trabalho_caso_autoctone = s.doenca_rel_trabalho_caso_autoctone
			else:
				doenca_rel_trabalho_caso_autoctone = ""

			if s.evolucao_caso != None:
				evolucao_caso = s.evolucao_caso
			else:
				evolucao_caso = ""

			if s.data_obito != None:
				data_obito = s.data_obito
			else:
				data_obito = ""

			if s.data_encerramento != None:
				data_encerramento = s.data_encerramento
			else:
				data_encerramento = ""

			if s.observacao != None:
				observacao = s.observacao
			else:
				observacao = ""

			if s.nome_investigador != None:
				nome_investigador = s.nome_investigador
			else:
				nome_investigador = ""

			if s.funcao_investigador != None:
				funcao_investigador = s.funcao_investigador
			else:
				funcao_investigador = ""

			if s.email_investigador != None:
				email_investigador = s.email_investigador
			else:
				email_investigador = ""

			if s.telefone_investigador != None:
				telefone_investigador = s.telefone_investigador
			else:
				telefone_investigador = ""

			if s.conselho_classe_investigador != None:
				conselho_classe_investigador = s.conselho_classe_investigador
			else:
				conselho_classe_investigador = ""

			if s.numero_unico != None:
				numero_unico = s.numero_unico
			else:
				numero_unico = ""

		

		row = [
			tipo_notificacao,
			agravo_doenca,
			codigo_cib10,
			data_notificacao,
			estado,
			municipio,
			codigo_ibge,
			data_primeiros_sintomas,
			unidade_saude,
			unidade_saude_outro,
			nome_paciente,
			data_nascimento_paciente,
			idade_paciente,
			sexo_paciente,
			paciente_gestante,
			raca_paciente,
			escolaridade_paciente,
			cartao_sus_paciente,
			nome_mae_paciente,
			cep_residencia,
			uf_residencia,
			municipio_residencia,
			bairro_residencia,
			codigo_ibge_residencia,
			rua_residencia,
			numero_residencia,
			complemento_residencia,
			distrito_residencia,
			ponto_referencia_residencia,
			telefone_residencia,
			zona_residencia,
			pais_residencia,
			data_investigacao,
			ocupacao,
			ambientes_frequentados,
			animais_que_teve_contato,
			natureza_contato_animais,
			relacao_animal_doente,
			exerce_atividade_contato_plantas,
			historico_contato_material,
			presenca_lesao_pele,
			natureza_lesao,
			natureza_lesao_outro,
			local_lesao,
			local_lesao_outro,
			diagnostico_forma_extrac_doenca,
			localizacao_forma_extrac_doenca,
			houve_coleta_material,
			data_coleta1,
			numero_gal1,
			data_coleta2,
			numero_gal2,
			data_coleta3,
			numero_gal3,
			resultado_isolamento,
			agente,
			histopatologia,
			data_resultado_exame1,
			descricao_exame_1,
			resultado_exame1,
			data_resultado_exame2,
			descricao_exame_2,
			resultado_exame2,
			data_resultado_exame3,
			descricao_exame_3,
			resultado_exame3,
			data_inicio_tratamento1,
			droga_administrada1,
			esquema_terapeutico1,
			data_inicio_tratamento2,
			droga_administrada2,
			esquema_terapeutico2,
			data_inicio_tratamento3,
			droga_administrada3,
			esquema_terapeutico3,
			hospitalizacao,
			data_internacao,
			data_da_alta,
			uf_hospitalizacao,
			municipio_hospitalizacao,
			codigo_ibge_hospitalizacao,
			nome_hospital_hospitalizacao,
			classificacao_final,
			criterio_confirmacao,
			caso_autoctone_municipio_residencia,
			uf_caso_autoctone,
			pais_caso_autoctone,
			municipio_caso_autoctone,
			codigo_ibge_caso_autoctone,
			distrito_caso_autoctone,
			bairro_caso_autoctone,
			area_provavel_infeccao_caso_autoctone,
			ambiente_infeccao_caso_autoctone,
			doenca_rel_trabalho_caso_autoctone,
			evolucao_caso,
			data_obito,
			data_encerramento,
			observacao,
			nome_investigador,
			funcao_investigador,
			email_investigador,
			telefone_investigador,
			conselho_classe_investigador,
			numero_unico,

		]
		for col_num, cell_value in enumerate(row, 1):
			cell = worksheet.cell(row=row_num, column=col_num)
			cell_value_ = str(cell_value)
			cell.value = cell_value_

		workbook.save(response)

		return response


	'''

'''


	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename="geral.csv"'

	writer = csv.writer(response)
	writer.writerow([
		'id','tipo_notificacao','agravo_doenca','codigo_cib10','data_notificacao','estado','municipio','codigo_ibge',
		'data_primeiros_sintomas','unidade_saude','unidade_saude_outro','nome_paciente','data_nascimento_paciente',
		'idade_paciente','sexo_paciente','paciente_gestante','raca_paciente','escolaridade_paciente','cartao_sus_paciente',
		'nome_mae_paciente','cep_residencia','uf_residencia','municipio_residencia','bairro_residencia',
		'codigo_ibge_residencia','rua_residencia','numero_residencia','complemento_residencia','distrito_residencia',
		'ponto_referencia_residencia','telefone_residencia','zona_residencia','pais_residencia','data_investigacao',
		'ocupacao','ambientes_frequentados','animais_que_teve_contato','natureza_contato_animais','relacao_animal_doente',
		'exerce_atividade_contato_plantas','historico_contato_material','presenca_lesao_pele','natureza_lesao',
		'natureza_lesao_outro','local_lesao','local_lesao_outro','diagnostico_forma_extrac_doenca',
		'localizacao_forma_extrac_doenca','houve_coleta_material','data_coleta1','numero_gal1','data_coleta2',
		'numero_gal2','data_coleta3','numero_gal3','resultado_isolamento','agente','histopatologia','data_resultado_exame1',
		'descricao_exame_1','resultado_exame1','data_resultado_exame2','descricao_exame_2','resultado_exame2',
		'data_resultado_exame3','descricao_exame_3','resultado_exame3','data_inicio_tratamento1','droga_administrada1',
		'esquema_terapeutico1','data_inicio_tratamento2','droga_administrada2','esquema_terapeutico2',
		'data_inicio_tratamento3','droga_administrada3','esquema_terapeutico3','hospitalizacao','data_internacao',
		'data_da_alta','uf_hospitalizacao','municipio_hospitalizacao','codigo_ibge_hospitalizacao',
		'nome_hospital_hospitalizacao','classificacao_final','criterio_confirmacao','caso_autoctone_municipio_residencia',
		'uf_caso_autoctone','pais_caso_autoctone','municipio_caso_autoctone','codigo_ibge_caso_autoctone',
		'distrito_caso_autoctone','bairro_caso_autoctone','area_provavel_infeccao_caso_autoctone',
		'ambiente_infeccao_caso_autoctone','doenca_rel_trabalho_caso_autoctone','evolucao_caso',
		'data_obito','data_encerramento','observacao','nome_investigador','funcao_investigador','email_investigador',
		'telefone_investigador','conselho_classe_investigador','numero_unico','gerencia_id',
		'responsavel_pelas_informacoes_id'

		])
'''


@login_required(login_url='/login/')
def organograma(request):
	return render(request, 'organograma.html')


@login_required(login_url='/login/')
def principal(request):	
	print(request.user.funcao)
	return render(request, 'principal.html')

@login_required(login_url='/login/')
def export_users(request):
	if request.user.funcao == 'admin':
		
		Users = get_user_model().objects.all().order_by('id')
		lista_users =[]
		for i in Users:
			user_id = i.id
			funcao = i.funcao
			login = i.login
			username = i.username
			cpf = i.cpf
			telefone = i.telefone
			unidade_saude = i.unidade_saude
			perfil = i.perfil
			email = i.email
			nome = i.first_name
			is_superuser = i.is_superuser
			is_staff = i.is_staff
			is_active = i.is_active
			last_login = i.last_login
			date_joined = i.date_joined
			gerencia_operacional = i.gerencia_operacional
			nucleo = i.nucleo
			area_tecnica = i.area_tecnica
			gerencia_regional = i.gerencia_regional
			municipio_nome = i.municipio_nome
			municipio_id = i.municipio_id
			lista_users.append({
				'user_id':user_id,
				'funcao':funcao,
				'login':login,
				'username':username,
				'cpf':cpf,
				'telefone':telefone,
				'unidade_saude':unidade_saude,
				'perfil':perfil,
				'email':email,
				'nome':nome,
				'is_superuser':is_superuser,
				'is_staff':is_staff,
				'is_active':is_active,
				'last_login':last_login,
				'date_joined':date_joined,
				'gerencia_operacional':gerencia_operacional,
				'nucleo':nucleo,
				'area_tecnica':area_tecnica,
				'gerencia_regional':gerencia_regional,
				'municipio_nome':municipio_nome,
				'municipio_id':municipio_id,
				})	
		
		df = pd.DataFrame(lista_users)
		try:
			for i in df['municipio_id']:
				
				try: # Checando se o valor é diferente de NaN
					i = int(i)
				except:
					continue
				else: # Buscando no modelo municipio o nome de municipio pelo id e alterando o dataframe
					municipio = Municipio.objects.get(id=i)
					df['municipio_nome'] = municipio.nome

			df['last_login'] = df['last_login'].apply(lambda a: pd.to_datetime(a).date())
			df['date_joined'] = df['date_joined'].apply(lambda a: pd.to_datetime(a).date())
			# Escrevendo o excel e enviando o response.
		except:
			redirect('principal')
		else:
			with BytesIO() as b:
				
				writer = pd.ExcelWriter(b, engine='openpyxl')
				df.to_excel(writer, sheet_name='Sheet1', index=False)
				writer.save()
				
				filename = 'Usuarios-SISGEVS.xlsx'
				response = HttpResponse(b.getvalue(),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
				response['Content-Disposition'] = 'attachment; filename=%s' % filename
				
				return response
	else:
		redirect('principal')
